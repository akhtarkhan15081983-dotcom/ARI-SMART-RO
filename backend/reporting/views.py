import calendar
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from django.db.models import Count, DecimalField, ExpressionWrapper, F, Q, Sum
from django.http import HttpResponse
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from accounts.permissions import IsAdminOrManager
from attendance.models import Attendance
from complaints.models import Complaint
from customers.models import Customer, CustomerRentHistory, CustomerRentPayment
from installation.models import Installation
from inventory.models import EngineerBagItem, InventoryItem, PartRequest
from jobs.models import Job, JobPartUsed
from purchase.models import PurchaseItem
from service.models import Service, ServicePart
from employees.models import EmployeeProfile, Holiday, LeaveRequest, PayrollRecord

from .periods import resolve_period


ZERO = Decimal("0.00")


def _decimal(value):
    return value if isinstance(value, Decimal) else Decimal(str(value or 0))


def _money(value):
    return f"{_decimal(value):.2f}"


def _datetime_filter(period, field):
    return {
        f"{field}__gte": period.start_datetime,
        f"{field}__lt": period.end_datetime,
    }


def _status_counts(queryset, field="status"):
    return {
        str(row[field] or "UNKNOWN"): row["count"]
        for row in queryset.values(field).annotate(count=Count("id")).order_by(field)
    }


def _employee_name(row, prefix):
    first_name = str(row.get(f"{prefix}__user__first_name") or "").strip()
    last_name = str(row.get(f"{prefix}__user__last_name") or "").strip()
    phone = str(row.get(f"{prefix}__user__phone") or "").strip()
    return f"{first_name} {last_name}".strip() or phone or "Unassigned"


def build_parts_report(period):
    usage = JobPartUsed.objects.filter(
        **_datetime_filter(period, "used_at")
    ).select_related(
        "job__engineer__user",
        "inventory_item__part",
    )

    totals = usage.aggregate(
        entries=Count("id"),
        quantity=Sum("quantity"),
        jobs=Count("job", distinct=True),
        employees=Count("job__engineer", distinct=True),
        parts=Count("inventory_item__part", distinct=True),
    )

    by_employee_rows = usage.values(
        "job__engineer__employee_id",
        "job__engineer__user__first_name",
        "job__engineer__user__last_name",
        "job__engineer__user__phone",
    ).annotate(
        usage_entries=Count("id"),
        quantity=Sum("quantity"),
        jobs=Count("job", distinct=True),
        different_parts=Count("inventory_item__part", distinct=True),
    ).order_by("-quantity", "job__engineer__employee_id")

    by_employee = [
        {
            "employee_id": row["job__engineer__employee_id"],
            "employee_name": _employee_name(row, "job__engineer"),
            "usage_entries": row["usage_entries"],
            "quantity": row["quantity"] or 0,
            "jobs": row["jobs"],
            "different_parts": row["different_parts"],
        }
        for row in by_employee_rows
    ]

    by_part = list(
        usage.values(
            "inventory_item__part__id",
            "inventory_item__part__code",
            "inventory_item__part__name",
        ).annotate(
            usage_entries=Count("id"),
            quantity=Sum("quantity"),
            employees=Count("job__engineer", distinct=True),
            jobs=Count("job", distinct=True),
        ).order_by("-quantity", "inventory_item__part__name")
    )

    recent_usage = [
        {
            "used_at": item.used_at.isoformat(),
            "job_id": item.job.job_id,
            "employee_id": item.job.engineer.employee_id,
            "employee_name": item.job.engineer.user.get_full_name()
            or item.job.engineer.user.phone,
            "part_code": item.inventory_item.part.code,
            "part_name": item.inventory_item.part.name,
            "serial_number": item.inventory_item.serial_number or "",
            "quantity": item.quantity,
            "remarks": item.remarks,
        }
        for item in usage.order_by("-used_at")[:100]
    ]

    bag_issued = EngineerBagItem.objects.filter(
        **_datetime_filter(period, "issue_date")
    ).count()
    bag_installed = EngineerBagItem.objects.filter(
        install_date__isnull=False,
        **_datetime_filter(period, "install_date"),
    ).count()
    bag_returned = EngineerBagItem.objects.filter(
        return_date__isnull=False,
        **_datetime_filter(period, "return_date"),
    ).count()

    requests = PartRequest.objects.filter(
        **_datetime_filter(period, "created_at")
    )

    return {
        "summary": {
            "usage_entries": totals["entries"] or 0,
            "used_quantity": totals["quantity"] or 0,
            "jobs": totals["jobs"] or 0,
            "employees": totals["employees"] or 0,
            "different_parts": totals["parts"] or 0,
            "bag_items_issued": bag_issued,
            "bag_items_installed": bag_installed,
            "bag_items_returned": bag_returned,
            "part_requests": requests.count(),
            "requested_quantity": requests.aggregate(total=Sum("quantity"))["total"] or 0,
        },
        "part_request_status": _status_counts(requests),
        "inventory_snapshot": _status_counts(InventoryItem.objects.all()),
        "by_employee": by_employee,
        "by_part": by_part,
        "recent_usage": recent_usage,
        "recent_usage_limited_to": 100,
    }


def build_rent_report(period):
    history = CustomerRentHistory.objects.filter(
        rent_month__range=(period.start, period.end)
    ).select_related("customer")

    totals = history.aggregate(
        expected=Sum("expected_rent"),
        paid=Sum("paid_amount"),
        customers=Count("customer", distinct=True),
        rent_rows=Count("id"),
    )
    expected = _decimal(totals["expected"])
    paid = _decimal(totals["paid"])

    overdue_rows = history.values(
        "customer__id",
        "customer__customer_id",
        "customer__name",
        "customer__phone",
    ).annotate(
        expected=Sum("expected_rent"),
        paid=Sum("paid_amount"),
    ).order_by("customer__name")

    outstanding_customers = []
    for row in overdue_rows:
        due = max(_decimal(row["expected"]) - _decimal(row["paid"]), ZERO)
        if due > 0:
            outstanding_customers.append({
                "customer_id": row["customer__id"],
                "customer_code": row["customer__customer_id"],
                "customer_name": row["customer__name"],
                "phone": row["customer__phone"],
                "expected": _money(row["expected"]),
                "paid": _money(row["paid"]),
                "outstanding": _money(due),
            })
    outstanding_customers.sort(
        key=lambda row: Decimal(row["outstanding"]),
        reverse=True,
    )

    payments = CustomerRentPayment.objects.filter(
        payment_date__range=(period.start, period.end)
    ).select_related(
        "customer",
        "collected_by__user",
    )
    payment_total = _decimal(payments.aggregate(total=Sum("amount"))["total"])

    by_mode = [
        {
            "payment_mode": row["payment_mode"],
            "payments": row["payments"],
            "amount": _money(row["amount"]),
        }
        for row in payments.values("payment_mode").annotate(
            payments=Count("id"),
            amount=Sum("amount"),
        ).order_by("-amount")
    ]

    collector_rows = payments.values(
        "collected_by__employee_id",
        "collected_by__user__first_name",
        "collected_by__user__last_name",
        "collected_by__user__phone",
    ).annotate(
        payments=Count("id"),
        customers=Count("customer", distinct=True),
        amount=Sum("amount"),
    ).order_by("-amount")

    by_collector = [
        {
            "employee_id": row["collected_by__employee_id"],
            "employee_name": _employee_name(row, "collected_by"),
            "payments": row["payments"],
            "customers": row["customers"],
            "amount": _money(row["amount"]),
        }
        for row in collector_rows
    ]

    efficiency = float((paid / expected * 100).quantize(Decimal("0.01"))) if expected else 0.0

    return {
        "summary": {
            "rent_rows": totals["rent_rows"] or 0,
            "customers": totals["customers"] or 0,
            "expected": _money(expected),
            "paid": _money(paid),
            "outstanding": _money(max(expected - paid, ZERO)),
            "collection_efficiency_percent": efficiency,
            "payment_transactions": payments.count(),
            "payments_received": _money(payment_total),
            "customers_with_due": len(outstanding_customers),
        },
        "by_payment_mode": by_mode,
        "by_collector": by_collector,
        "outstanding_customers": outstanding_customers[:200],
        "outstanding_customers_limited_to": 200,
    }


def build_attendance_report(period):
    attendance = Attendance.objects.filter(
        date__range=(period.start, period.end)
    ).select_related("employee__user")

    totals = attendance.aggregate(
        records=Count("id"),
        employees=Count("employee", distinct=True),
        present=Count("id", filter=Q(status="PRESENT")),
        absent=Count("id", filter=Q(status="ABSENT")),
        half_day=Count("id", filter=Q(status="HALF_DAY")),
        leave=Count("id", filter=Q(status="LEAVE")),
        working_hours=Sum("working_hours"),
        checked_in=Count("id", filter=Q(check_in__isnull=False)),
        checked_out=Count("id", filter=Q(check_out__isnull=False)),
        pending_reviews=Count("id", filter=Q(identity_review_status="PENDING")),
        rejected_reviews=Count("id", filter=Q(identity_review_status="REJECTED")),
    )

    employee_rows = attendance.values(
        "employee__employee_id",
        "employee__user__first_name",
        "employee__user__last_name",
        "employee__user__phone",
        "employee__designation",
    ).annotate(
        records=Count("id"),
        present=Count("id", filter=Q(status="PRESENT")),
        absent=Count("id", filter=Q(status="ABSENT")),
        half_day=Count("id", filter=Q(status="HALF_DAY")),
        leave=Count("id", filter=Q(status="LEAVE")),
        working_hours=Sum("working_hours"),
        late_checkout_missing=Count(
            "id",
            filter=Q(check_in__isnull=False, check_out__isnull=True),
        ),
    ).order_by("employee__employee_id")

    by_employee = [
        {
            "employee_id": row["employee__employee_id"],
            "employee_name": _employee_name(row, "employee"),
            "designation": row["employee__designation"],
            "records": row["records"],
            "present": row["present"],
            "absent": row["absent"],
            "half_day": row["half_day"],
            "leave": row["leave"],
            "working_hours": _money(row["working_hours"]),
            "missing_checkout": row["late_checkout_missing"],
        }
        for row in employee_rows
    ]

    return {
        "summary": {
            "records": totals["records"] or 0,
            "employees": totals["employees"] or 0,
            "present": totals["present"] or 0,
            "absent": totals["absent"] or 0,
            "half_day": totals["half_day"] or 0,
            "leave": totals["leave"] or 0,
            "working_hours": _money(totals["working_hours"]),
            "checked_in": totals["checked_in"] or 0,
            "checked_out": totals["checked_out"] or 0,
            "pending_identity_reviews": totals["pending_reviews"] or 0,
            "rejected_identity_reviews": totals["rejected_reviews"] or 0,
        },
        "by_employee": by_employee,
    }


def build_operations_report(period):
    jobs_created = Job.objects.filter(**_datetime_filter(period, "created_at"))
    jobs_completed = Job.objects.filter(
        completed_at__isnull=False,
        **_datetime_filter(period, "completed_at"),
    )
    complaints_created = Complaint.objects.filter(
        **_datetime_filter(period, "complaint_date")
    )
    complaints_resolved = Complaint.objects.filter(
        resolved_date__isnull=False,
        **_datetime_filter(period, "resolved_date"),
    )
    installations = Installation.objects.filter(
        **_datetime_filter(period, "scheduled_date")
    )
    services = Service.objects.filter(
        **_datetime_filter(period, "scheduled_date")
    )
    new_customers = Customer.objects.filter(
        **_datetime_filter(period, "created_at")
    )

    purchase_value_expression = ExpressionWrapper(
        F("quantity") * F("purchase_price"),
        output_field=DecimalField(max_digits=18, decimal_places=2),
    )
    purchase_items = PurchaseItem.objects.filter(
        purchase__invoice_date__range=(period.start, period.end)
    )
    purchase_totals = purchase_items.aggregate(
        items=Count("id"),
        total_quantity=Sum("quantity"),
        value=Sum(purchase_value_expression),
        invoices=Count("purchase", distinct=True),
        suppliers=Count("purchase__supplier", distinct=True),
    )

    open_complaints = Complaint.objects.exclude(
        status__in=["RESOLVED", "CLOSED", "CANCELLED"]
    ).count()

    return {
        "summary": {
            "new_customers": new_customers.count(),
            "jobs_created": jobs_created.count(),
            "jobs_completed": jobs_completed.count(),
            "complaints_created": complaints_created.count(),
            "complaints_resolved": complaints_resolved.count(),
            "open_complaints_snapshot": open_complaints,
            "installations_scheduled": installations.count(),
            "services_scheduled": services.count(),
            "purchase_invoices": purchase_totals["invoices"] or 0,
            "purchase_suppliers": purchase_totals["suppliers"] or 0,
            "purchase_items": purchase_totals["items"] or 0,
            "purchase_quantity": purchase_totals["total_quantity"] or 0,
            "purchase_value": _money(purchase_totals["value"]),
        },
        "jobs_by_status": _status_counts(jobs_created),
        "jobs_by_type": _status_counts(jobs_created, "job_type"),
        "complaints_by_status": _status_counts(complaints_created),
        "complaints_by_type": _status_counts(complaints_created, "complaint_type"),
        "installations_by_status": _status_counts(installations),
        "services_by_status": _status_counts(services),
    }


def build_customer_report(period):
    """Customer-wise financial and operational audit for the selected period."""
    rows = {}

    def customer_row(customer_id, code, name, phone):
        return rows.setdefault(customer_id, {
            "customer_code": code,
            "customer_name": name,
            "phone": phone,
            "payments": 0,
            "payment_amount": ZERO,
            "parts_quantity": 0,
            "parts_types": 0,
            "jobs": 0,
            "jobs_completed": 0,
            "services": 0,
            "services_completed": 0,
            "complaints": 0,
            "complaints_resolved": 0,
        })

    payment_rows = CustomerRentPayment.objects.filter(
        payment_date__range=(period.start, period.end)
    ).values(
        "customer_id", "customer__customer_id", "customer__name", "customer__phone"
    ).annotate(
        payments=Count("id"), amount=Sum("amount")
    )
    for item in payment_rows:
        row = customer_row(item["customer_id"], item["customer__customer_id"], item["customer__name"], item["customer__phone"])
        row["payments"] = item["payments"]
        row["payment_amount"] = _decimal(item["amount"])

    part_rows = JobPartUsed.objects.filter(
        **_datetime_filter(period, "used_at")
    ).values(
        "job__customer_id", "job__customer__customer_id", "job__customer__name", "job__customer__phone"
    ).annotate(
        quantity=Sum("quantity"), part_types=Count("inventory_item__part", distinct=True)
    )
    for item in part_rows:
        row = customer_row(item["job__customer_id"], item["job__customer__customer_id"], item["job__customer__name"], item["job__customer__phone"])
        row["parts_quantity"] = item["quantity"] or 0
        row["parts_types"] = item["part_types"] or 0

    service_part_rows = ServicePart.objects.filter(
        **_datetime_filter(period, "service__scheduled_date")
    ).values(
        "service__customer_id", "service__customer__customer_id",
        "service__customer__name", "service__customer__phone",
    ).annotate(
        quantity=Sum("quantity"), part_types=Count("part", distinct=True)
    )
    for item in service_part_rows:
        row = customer_row(item["service__customer_id"], item["service__customer__customer_id"], item["service__customer__name"], item["service__customer__phone"])
        row["parts_quantity"] += item["quantity"] or 0
        row["parts_types"] += item["part_types"] or 0

    job_rows = Job.objects.filter(
        **_datetime_filter(period, "created_at")
    ).values(
        "customer_id", "customer__customer_id", "customer__name", "customer__phone"
    ).annotate(
        jobs=Count("id"), completed=Count("id", filter=Q(status="COMPLETED"))
    )
    for item in job_rows:
        row = customer_row(item["customer_id"], item["customer__customer_id"], item["customer__name"], item["customer__phone"])
        row["jobs"] = item["jobs"]
        row["jobs_completed"] = item["completed"]

    service_rows = Service.objects.filter(
        **_datetime_filter(period, "scheduled_date")
    ).values(
        "customer_id", "customer__customer_id", "customer__name", "customer__phone"
    ).annotate(
        services=Count("id"), completed=Count("id", filter=Q(status="COMPLETED"))
    )
    for item in service_rows:
        row = customer_row(item["customer_id"], item["customer__customer_id"], item["customer__name"], item["customer__phone"])
        row["services"] = item["services"]
        row["services_completed"] = item["completed"]

    complaint_rows = Complaint.objects.filter(
        **_datetime_filter(period, "complaint_date")
    ).values(
        "customer_id", "customer__customer_id", "customer__name", "customer__phone"
    ).annotate(
        complaints=Count("id"),
        resolved=Count("id", filter=Q(status__in=["RESOLVED", "CLOSED"])),
    )
    for item in complaint_rows:
        row = customer_row(item["customer_id"], item["customer__customer_id"], item["customer__name"], item["customer__phone"])
        row["complaints"] = item["complaints"]
        row["complaints_resolved"] = item["resolved"]

    summary = []
    for row in rows.values():
        row["payment_amount"] = _money(row["payment_amount"])
        summary.append(row)
    summary.sort(key=lambda row: (row["customer_name"].lower(), row["customer_code"]))

    payments = [{
        "payment_date": item["payment_date"],
        "customer_code": item["customer__customer_id"],
        "customer_name": item["customer__name"],
        "amount": _money(item["amount"]),
        "payment_mode": item["payment_mode"],
        "collected_by": item["collected_by__employee_id"] or "",
        "remarks": item["remarks"],
    } for item in CustomerRentPayment.objects.filter(
        payment_date__range=(period.start, period.end)
    ).values(
        "payment_date", "customer__customer_id", "customer__name", "amount",
        "payment_mode", "remarks", "collected_by__employee_id",
    ).order_by("customer__name", "-payment_date")]

    parts = [{
        "activity_date": item["used_at"],
        "customer_code": item["job__customer__customer_id"],
        "customer_name": item["job__customer__name"],
        "work_id": item["job__job_id"],
        "part_code": item["inventory_item__part__code"],
        "part_name": item["inventory_item__part__name"],
        "serial_number": item["inventory_item__serial_number"],
        "quantity": item["quantity"],
        "remarks": item["remarks"],
    } for item in JobPartUsed.objects.filter(
        **_datetime_filter(period, "used_at")
    ).values(
        "used_at", "job__customer__customer_id", "job__customer__name",
        "job__job_id", "inventory_item__part__code", "inventory_item__part__name",
        "inventory_item__serial_number", "quantity", "remarks",
    ).order_by("job__customer__name", "-used_at")]
    parts.extend({
        "activity_date": item["service__scheduled_date"],
        "customer_code": item["service__customer__customer_id"],
        "customer_name": item["service__customer__name"],
        "work_id": item["service__service_id"],
        "part_code": item["part__code"],
        "part_name": item["part__name"],
        "serial_number": item["inventory_item__serial_number"],
        "quantity": item["quantity"],
        "remarks": item["remarks"],
    } for item in ServicePart.objects.filter(
        **_datetime_filter(period, "service__scheduled_date")
    ).values(
        "service__scheduled_date", "service__customer__customer_id",
        "service__customer__name", "service__service_id", "part__code",
        "part__name", "inventory_item__serial_number", "quantity", "remarks",
    ))

    return {
        "summary": {
            "customers": len(summary),
            "customers_with_payments": sum(1 for row in summary if row["payments"] > 0),
            "customers_with_parts": sum(1 for row in summary if row["parts_quantity"] > 0),
        },
        "by_customer": summary,
        "payment_details": payments,
        "part_details": parts,
    }


def build_management_report(period, parts, rent, attendance, operations):
    rent_summary = rent["summary"]
    operation_summary = operations["summary"]
    attendance_summary = attendance["summary"]
    jobs_created = operation_summary["jobs_created"] or 0
    jobs_completed = operation_summary["jobs_completed"] or 0
    complaints_created = operation_summary["complaints_created"] or 0
    complaints_resolved = operation_summary["complaints_resolved"] or 0
    attendance_records = attendance_summary["records"] or 0
    present = attendance_summary["present"] or 0
    collection_rate = float(rent_summary["collection_efficiency_percent"] or 0)
    job_rate = round(jobs_completed / jobs_created * 100, 1) if jobs_created else 0
    complaint_rate = round(complaints_resolved / complaints_created * 100, 1) if complaints_created else 0
    attendance_rate = round(present / attendance_records * 100, 1) if attendance_records else 0
    scored_rates = [rate for rate, count in [
        (collection_rate, rent_summary["rent_rows"]),
        (job_rate, jobs_created),
        (complaint_rate, complaints_created),
        (attendance_rate, attendance_records),
    ] if count]
    health_score = round(sum(scored_rates) / len(scored_rates), 1) if scored_rates else 0

    ageing = {"current": ZERO, "31_60_days": ZERO, "61_90_days": ZERO, "over_90_days": ZERO}
    for row in CustomerRentHistory.objects.filter(
        rent_month__lte=period.end
    ).values("rent_month", "expected_rent", "paid_amount"):
        due = max(_decimal(row["expected_rent"]) - _decimal(row["paid_amount"]), ZERO)
        if not due:
            continue
        age = (period.end - row["rent_month"]).days
        bucket = "current" if age <= 30 else "31_60_days" if age <= 60 else "61_90_days" if age <= 90 else "over_90_days"
        ageing[bucket] += due

    trend = []
    anchor = period.end.replace(day=1)
    for offset in range(5, -1, -1):
        month_index = anchor.year * 12 + anchor.month - 1 - offset
        year, month_zero = divmod(month_index, 12)
        month = month_zero + 1
        start = date(year, month, 1)
        end = date(year, month, calendar.monthrange(year, month)[1])
        payments = _decimal(CustomerRentPayment.objects.filter(
            payment_date__range=(start, end)
        ).aggregate(total=Sum("amount"))["total"])
        completed = Job.objects.filter(completed_at__date__range=(start, end)).count()
        resolved = Complaint.objects.filter(resolved_date__date__range=(start, end)).count()
        trend.append({"month": start.strftime("%b %Y"), "rent_collected": _money(payments), "jobs_completed": completed, "complaints_resolved": resolved})

    employee_rows = {}
    for row in rent["by_collector"]:
        key = row["employee_id"] or "UNASSIGNED"
        employee_rows[key] = {"employee_id": key, "employee_name": row["employee_name"], "rent_collected": row["amount"], "payments": row["payments"], "parts_used": 0, "jobs": 0, "working_hours": "0.00"}
    for row in parts["by_employee"]:
        key = row["employee_id"] or "UNASSIGNED"
        target = employee_rows.setdefault(key, {"employee_id": key, "employee_name": row["employee_name"], "rent_collected": "0.00", "payments": 0, "parts_used": 0, "jobs": 0, "working_hours": "0.00"})
        target["parts_used"] = row["quantity"]
        target["jobs"] = row["jobs"]
    for row in attendance["by_employee"]:
        key = row["employee_id"] or "UNASSIGNED"
        target = employee_rows.setdefault(key, {"employee_id": key, "employee_name": row["employee_name"], "rent_collected": "0.00", "payments": 0, "parts_used": 0, "jobs": 0, "working_hours": "0.00"})
        target["working_hours"] = row["working_hours"]

    insights = []
    if collection_rate < 80:
        insights.append({"severity": "HIGH", "title": "Collection efficiency needs attention", "detail": f"Only {collection_rate:.1f}% of expected rent is collected; ₹{rent_summary['outstanding']} remains outstanding."})
    if operation_summary["open_complaints_snapshot"]:
        insights.append({"severity": "MEDIUM", "title": "Open complaint backlog", "detail": f"{operation_summary['open_complaints_snapshot']} complaints are currently open."})
    if attendance_summary["pending_identity_reviews"]:
        insights.append({"severity": "MEDIUM", "title": "Attendance reviews pending", "detail": f"{attendance_summary['pending_identity_reviews']} identity reviews require action."})
    if not insights:
        insights.append({"severity": "GOOD", "title": "No critical exception detected", "detail": "Core operating indicators are within a healthy range for this period."})

    return {
        "scorecard": {
            "business_health_score": health_score,
            "collection_efficiency": collection_rate,
            "job_completion_rate": job_rate,
            "complaint_resolution_rate": complaint_rate,
            "attendance_rate": attendance_rate,
        },
        "rent_ageing": {key: _money(value) for key, value in ageing.items()},
        "trend_6_months": trend,
        "employee_productivity": sorted(employee_rows.values(), key=lambda row: Decimal(str(row["rent_collected"] or 0)), reverse=True),
        "insights": insights,
    }


def build_reports(period):
    parts = build_parts_report(period)
    rent = build_rent_report(period)
    attendance = build_attendance_report(period)
    operations = build_operations_report(period)
    customers = build_customer_report(period)
    management = build_management_report(period, parts, rent, attendance, operations)
    hrms = build_hrms_report(period)
    ledgers = build_operational_ledgers(period)

    return {
        "period": period.as_dict(),
        "overview": {
            "parts_used": parts["summary"]["used_quantity"],
            "rent_collected": rent["summary"]["payments_received"],
            "rent_outstanding": rent["summary"]["outstanding"],
            "attendance_hours": attendance["summary"]["working_hours"],
            "jobs_completed": operations["summary"]["jobs_completed"],
            "open_complaints": operations["summary"]["open_complaints_snapshot"],
        },
        "parts": parts,
        "rent": rent,
        "attendance": attendance,
        "operations": operations,
        "customers": customers,
        "management": management,
        "hrms": hrms,
        "ledgers": ledgers,
    }


def build_operational_ledgers(period):
    def employee_name(employee):
        return employee.user.get_full_name() or employee.user.phone

    jobs = [{
        "job_id": row.job_id,
        "customer_code": row.customer.customer_id,
        "customer_name": row.customer.name,
        "customer_phone": row.customer.phone,
        "engineer_id": row.engineer.employee_id,
        "engineer_name": employee_name(row.engineer),
        "asset_id": row.ro_asset.asset_id,
        "job_type": row.get_job_type_display(),
        "priority": row.get_priority_display(),
        "scheduled_date": row.scheduled_date,
        "status": row.get_status_display(),
        "assigned_at": row.assigned_at,
        "completed_at": row.completed_at,
        "remarks": row.remarks,
    } for row in Job.objects.filter(
        **_datetime_filter(period, "scheduled_date")
    ).select_related("customer", "engineer__user", "ro_asset").order_by("scheduled_date")]

    complaints = [{
        "complaint_id": row.complaint_id,
        "customer_code": row.customer.customer_id,
        "customer_name": row.customer.name,
        "customer_phone": row.customer.phone,
        "engineer_id": row.engineer.employee_id if row.engineer else "",
        "engineer_name": employee_name(row.engineer) if row.engineer else "Unassigned",
        "complaint_type": row.get_complaint_type_display(),
        "priority": row.get_priority_display(),
        "complaint_date": row.complaint_date,
        "scheduled_date": row.scheduled_date,
        "resolved_date": row.resolved_date,
        "status": row.get_status_display(),
        "description": row.description,
        "resolution": row.resolution,
        "engineer_remarks": row.engineer_remarks,
    } for row in Complaint.objects.filter(
        **_datetime_filter(period, "complaint_date")
    ).select_related("customer", "engineer__user").order_by("complaint_date")]

    services = [{
        "service_id": row.service_id,
        "customer_code": row.customer.customer_id,
        "customer_name": row.customer.name,
        "customer_phone": row.customer.phone,
        "engineer_id": row.engineer.employee_id,
        "engineer_name": employee_name(row.engineer),
        "asset_id": row.ro_asset.asset_id,
        "service_type": row.get_service_type_display(),
        "scheduled_date": row.scheduled_date,
        "completed_date": row.completed_date,
        "input_tds": row.input_tds,
        "output_tds": row.output_tds,
        "next_service_date": row.next_service_date,
        "status": row.get_status_display(),
        "remarks": row.remarks,
    } for row in Service.objects.filter(
        **_datetime_filter(period, "scheduled_date")
    ).select_related("customer", "engineer__user", "ro_asset").order_by("scheduled_date")]

    installations = [{
        "installation_id": row.installation_id,
        "customer_code": row.customer.customer_id,
        "customer_name": row.customer.name,
        "customer_phone": row.customer.phone,
        "engineer_id": row.engineer.employee_id,
        "engineer_name": employee_name(row.engineer),
        "asset_id": row.ro_asset.asset_id,
        "business_type": row.get_business_type_display(),
        "scheduled_date": row.scheduled_date,
        "completed_date": row.completed_date,
        "input_tds": row.input_tds,
        "output_tds": row.output_tds,
        "referral_name": row.referral_name,
        "status": row.get_status_display(),
        "remarks": row.remarks,
    } for row in Installation.objects.filter(
        **_datetime_filter(period, "scheduled_date")
    ).select_related("customer", "engineer__user", "ro_asset").order_by("scheduled_date")]

    return {
        "jobs": jobs,
        "complaints": complaints,
        "services": services,
        "installations": installations,
    }


def build_hrms_report(period):
    employees = []
    for employee in EmployeeProfile.objects.select_related("user").order_by("employee_id"):
        employees.append({
            "employee_id": employee.employee_id,
            "employee_name": employee.user.get_full_name() or employee.user.phone,
            "phone": employee.user.phone,
            "designation": employee.get_designation_display(),
            "joining_date": employee.joining_date,
            "monthly_salary": employee.salary,
            "active": employee.is_active,
            "face_enrolled": bool(employee.face_enrolled_at),
            "attendance_device_bound": bool(employee.attendance_device_id),
        })

    leaves = [{
        "employee_id": row.employee.employee_id,
        "employee_name": row.employee.user.get_full_name() or row.employee.user.phone,
        "leave_type": row.get_leave_type_display(),
        "start_date": row.start_date,
        "end_date": row.end_date,
        "status": row.status,
        "paid": row.is_paid,
        "reason": row.reason,
        "review_note": row.review_note,
        "reviewed_by": (row.reviewed_by.get_full_name() or row.reviewed_by.phone) if row.reviewed_by else "",
    } for row in LeaveRequest.objects.filter(
        start_date__lte=period.end,
        end_date__gte=period.start,
    ).select_related("employee__user", "reviewed_by")]

    payroll = [{
        "employee_id": row.employee.employee_id,
        "employee_name": row.employee.user.get_full_name() or row.employee.user.phone,
        "payroll_month": row.payroll_month,
        "base_salary": row.base_salary,
        "payable_base": row.payable_base,
        "late_days": row.late_days,
        "late_penalty": row.late_penalty,
        "half_day_deduction": row.half_day_deduction,
        "absence_deduction": row.absence_deduction,
        "overtime_hours": row.overtime_hours,
        "overtime_amount": row.overtime_amount,
        "rent_incentive": row.rent_incentive,
        "sale_incentive": row.sale_incentive,
        "other_earnings": row.other_earnings,
        "other_deductions": row.other_deductions,
        "net_salary": row.net_salary,
        "status": row.status,
        "paid_at": row.paid_at,
    } for row in PayrollRecord.objects.filter(
        payroll_month__gte=period.start.replace(day=1),
        payroll_month__lte=period.end,
    ).select_related("employee__user")]

    holidays = [{
        "date": row.date,
        "holiday": row.name,
        "description": row.description,
        "paid": row.is_paid,
        "declared_by": row.declared_by.get_full_name() or row.declared_by.phone,
    } for row in Holiday.objects.filter(date__range=(period.start, period.end)).select_related("declared_by")]

    return {
        "summary": {
            "active_employees": sum(1 for row in employees if row["active"]),
            "pending_leaves": sum(1 for row in leaves if row["status"] == "PENDING"),
            "approved_leaves": sum(1 for row in leaves if row["status"] == "APPROVED"),
            "payroll_net": _money(sum((_decimal(row["net_salary"]) for row in payroll), ZERO)),
            "overtime_amount": _money(sum((_decimal(row["overtime_amount"]) for row in payroll), ZERO)),
            "incentives": _money(sum((_decimal(row["rent_incentive"]) + _decimal(row["sale_incentive"]) for row in payroll), ZERO)),
            "holidays": len(holidays),
        },
        "employees": employees,
        "leaves": leaves,
        "payroll": payroll,
        "holidays": holidays,
    }


def _append_rows(worksheet, rows):
    if not rows:
        worksheet.append(["No records"])
        return
    headers = list(rows[0].keys())
    worksheet.append([_humanize_header(header) for header in headers])
    for row in rows:
        worksheet.append([_excel_value(row.get(header, ""), header) for header in headers])


def _humanize_header(value):
    return str(value).replace("__", " ").replace("_", " ").title()


def _excel_value(value, field=""):
    if isinstance(value, datetime):
        return value.astimezone().strftime("%d-%m-%Y %I:%M %p")
    if isinstance(value, date):
        return value.strftime("%d-%m-%Y")
    numeric_fields = {
        "amount", "payment_amount", "rent_collected", "rent_outstanding",
        "expected", "paid", "outstanding", "purchase_value", "working_hours",
        "current", "31_60_days", "61_90_days", "over_90_days",
        "monthly_salary", "base_salary", "payable_base", "late_penalty",
        "half_day_deduction", "absence_deduction", "overtime_amount",
        "rent_incentive", "sale_incentive", "other_earnings",
        "other_deductions", "net_salary", "payroll_net", "incentives",
    }
    if field in numeric_fields and isinstance(value, str):
        try:
            return float(value)
        except ValueError:
            return value
    return value


def build_workbook(data):
    workbook = openpyxl.Workbook()
    dashboard = workbook.active
    dashboard.title = "Executive Dashboard"
    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet.append(["ARI SMART RO REPORT"])
    summary_sheet.append(["Period", data["period"]["label"]])
    summary_sheet.append([])
    summary_sheet.append(["Metric", "Value"])
    for key, value in data["overview"].items():
        summary_sheet.append([_humanize_header(key), _excel_value(value, key)])

    parts_sheet = workbook.create_sheet("Parts by Employee")
    _append_rows(parts_sheet, data["parts"]["by_employee"])

    part_sheet = workbook.create_sheet("Parts by Item")
    _append_rows(part_sheet, data["parts"]["by_part"])

    rent_sheet = workbook.create_sheet("Rent Due")
    _append_rows(rent_sheet, data["rent"]["outstanding_customers"])

    collector_sheet = workbook.create_sheet("Rent Collectors")
    _append_rows(collector_sheet, data["rent"]["by_collector"])

    attendance_sheet = workbook.create_sheet("Attendance")
    _append_rows(attendance_sheet, data["attendance"]["by_employee"])

    operations_sheet = workbook.create_sheet("Operations")
    operations_sheet.append(["Metric", "Value"])
    for key, value in data["operations"]["summary"].items():
        operations_sheet.append([_humanize_header(key), _excel_value(value, key)])

    customer_sheet = workbook.create_sheet("Customer Summary")
    _append_rows(customer_sheet, data["customers"]["by_customer"])

    customer_payment_sheet = workbook.create_sheet("Customer Payments")
    _append_rows(customer_payment_sheet, data["customers"]["payment_details"])

    customer_parts_sheet = workbook.create_sheet("Customer Parts")
    _append_rows(customer_parts_sheet, data["customers"]["part_details"])

    job_register_sheet = workbook.create_sheet("Job Register")
    _append_rows(job_register_sheet, data["ledgers"]["jobs"])

    complaint_register_sheet = workbook.create_sheet("Complaint Register")
    _append_rows(complaint_register_sheet, data["ledgers"]["complaints"])

    service_register_sheet = workbook.create_sheet("Service Register")
    _append_rows(service_register_sheet, data["ledgers"]["services"])

    installation_register_sheet = workbook.create_sheet("Installation Register")
    _append_rows(installation_register_sheet, data["ledgers"]["installations"])

    employee_productivity_sheet = workbook.create_sheet("Employee Productivity")
    _append_rows(employee_productivity_sheet, data["management"]["employee_productivity"])

    rent_ageing_sheet = workbook.create_sheet("Rent Ageing")
    _append_rows(rent_ageing_sheet, [data["management"]["rent_ageing"]])

    insight_sheet = workbook.create_sheet("Management Actions")
    _append_rows(insight_sheet, data["management"]["insights"])

    hr_summary_sheet = workbook.create_sheet("HR Executive Summary")
    hr_summary_sheet.append(["Metric", "Value"])
    for key, value in data["hrms"]["summary"].items():
        hr_summary_sheet.append([_humanize_header(key), _excel_value(value, key)])

    employee_master_sheet = workbook.create_sheet("HR Employee Master")
    _append_rows(employee_master_sheet, data["hrms"]["employees"])

    leave_sheet = workbook.create_sheet("HR Leave Register")
    _append_rows(leave_sheet, data["hrms"]["leaves"])

    payroll_sheet = workbook.create_sheet("HR Payroll Register")
    _append_rows(payroll_sheet, data["hrms"]["payroll"])

    holiday_sheet = workbook.create_sheet("HR Holiday Calendar")
    _append_rows(holiday_sheet, data["hrms"]["holidays"])

    index_sheet = workbook.create_sheet("Workbook Index", 1)
    index_sheet.append(["Sheet", "Purpose", "Records"])
    index_rows = [
        ("Executive Dashboard", "Leadership KPIs, trends and management actions", 1),
        ("Customer Summary", "Customer-wise business activity and balances", len(data["customers"]["by_customer"])),
        ("Customer Payments", "Customer-wise payment ledger", len(data["customers"]["payment_details"])),
        ("Customer Parts", "Customer-wise parts and service usage", len(data["customers"]["part_details"])),
        ("Job Register", "Complete customer and engineer work ledger", len(data["ledgers"]["jobs"])),
        ("Complaint Register", "Complaint lifecycle, ownership and resolution", len(data["ledgers"]["complaints"])),
        ("Service Register", "Service history, TDS readings and next service", len(data["ledgers"]["services"])),
        ("Installation Register", "Rental, sale and AMC installation history", len(data["ledgers"]["installations"])),
        ("Attendance", "Employee attendance and working-hours summary", len(data["attendance"]["by_employee"])),
        ("HR Employee Master", "Safe employee master without Aadhaar/PAN", len(data["hrms"]["employees"])),
        ("HR Leave Register", "Leave requests, approvals and review notes", len(data["hrms"]["leaves"])),
        ("HR Payroll Register", "Salary, penalties, overtime and incentives", len(data["hrms"]["payroll"])),
        ("HR Holiday Calendar", "Declared paid office holidays", len(data["hrms"]["holidays"])),
    ]
    for row in index_rows:
        index_sheet.append(row)

    navy = "123A63"
    blue = "0878C9"
    pale_blue = "EAF4FC"
    white = "FFFFFF"
    border = Border(bottom=Side(style="thin", color="D9E2EC"))

    management = data["management"]
    scorecard = management["scorecard"]
    dashboard.merge_cells("A1:L2")
    dashboard["A1"] = "ARI SMART RO  |  EXECUTIVE BUSINESS REVIEW"
    dashboard["A1"].font = Font(size=22, bold=True, color=white)
    dashboard["A1"].fill = PatternFill("solid", fgColor=navy)
    dashboard["A1"].alignment = Alignment(horizontal="left", vertical="center")
    dashboard.merge_cells("A3:L3")
    dashboard["A3"] = f"Reporting period: {data['period']['label']}  |  Management view"
    dashboard["A3"].font = Font(size=11, color="52606D")

    cards = [
        ("A5:C7", "Business Health", scorecard["business_health_score"], "0.0"),
        ("D5:F7", "Collection Efficiency", scorecard["collection_efficiency"] / 100, "0.0%"),
        ("G5:I7", "Job Completion", scorecard["job_completion_rate"] / 100, "0.0%"),
        ("J5:L7", "Complaint Resolution", scorecard["complaint_resolution_rate"] / 100, "0.0%"),
    ]
    for cell_range, label, value, number_format in cards:
        dashboard.merge_cells(cell_range)
        cell = dashboard[cell_range.split(":")[0]]
        cell.value = f"{label}\n{value:.1f}" if number_format == "0.0" else f"{label}\n{value:.1%}"
        cell.font = Font(size=14, bold=True, color=navy)
        cell.fill = PatternFill("solid", fgColor="EAF4FC")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    dashboard["A9"] = "MANAGEMENT EXCEPTIONS & ACTIONS"
    dashboard["A9"].font = Font(size=13, bold=True, color=white)
    dashboard["A9"].fill = PatternFill("solid", fgColor=blue)
    dashboard.merge_cells("A9:L9")
    for index, insight in enumerate(management["insights"], start=10):
        dashboard.merge_cells(start_row=index, start_column=1, end_row=index, end_column=12)
        cell = dashboard.cell(index, 1)
        cell.value = f"[{insight['severity']}] {insight['title']} — {insight['detail']}"
        cell.fill = PatternFill("solid", fgColor="FFF4E5" if insight["severity"] != "GOOD" else "E8F5E9")
        cell.font = Font(color="9C2C13" if insight["severity"] == "HIGH" else navy, bold=insight["severity"] == "HIGH")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        dashboard.row_dimensions[index].height = 30

    trend_start = 31
    dashboard.cell(trend_start, 1, "Month")
    dashboard.cell(trend_start, 2, "Rent Collected")
    dashboard.cell(trend_start, 3, "Jobs Completed")
    dashboard.cell(trend_start, 4, "Complaints Resolved")
    for row_index, row in enumerate(management["trend_6_months"], start=trend_start + 1):
        dashboard.cell(row_index, 1, row["month"])
        dashboard.cell(row_index, 2, float(row["rent_collected"]))
        dashboard.cell(row_index, 3, row["jobs_completed"])
        dashboard.cell(row_index, 4, row["complaints_resolved"])
    trend_end = trend_start + len(management["trend_6_months"])

    rent_chart = LineChart()
    rent_chart.title = "Rent Collection Trend (₹)"
    rent_chart.style = 13
    rent_chart.height = 7.5
    rent_chart.width = 12.5
    rent_chart.add_data(Reference(dashboard, min_col=2, min_row=trend_start, max_row=trend_end), titles_from_data=True)
    rent_chart.set_categories(Reference(dashboard, min_col=1, min_row=trend_start + 1, max_row=trend_end))
    rent_chart.y_axis.numFmt = '₹#,##0'
    dashboard.add_chart(rent_chart, "A14")

    work_chart = BarChart()
    work_chart.type = "col"
    work_chart.title = "Completed Work Trend"
    work_chart.style = 10
    work_chart.height = 7.5
    work_chart.width = 12.5
    work_chart.add_data(Reference(dashboard, min_col=3, max_col=4, min_row=trend_start, max_row=trend_end), titles_from_data=True)
    work_chart.set_categories(Reference(dashboard, min_col=1, min_row=trend_start + 1, max_row=trend_end))
    dashboard.add_chart(work_chart, "G14")

    dashboard.sheet_state = "visible"
    dashboard.sheet_view.showGridLines = False
    dashboard.freeze_panes = "A4"
    dashboard.column_dimensions["A"].width = 16
    for column in range(2, 13):
        dashboard.column_dimensions[get_column_letter(column)].width = 12
    dashboard.row_dimensions[1].height = 28
    dashboard.sheet_properties.pageSetUpPr.fitToPage = True
    dashboard.page_setup.orientation = "landscape"
    dashboard.page_setup.fitToWidth = 1

    summary_sheet.merge_cells("A1:B1")
    summary_sheet["A1"].font = Font(size=18, bold=True, color=white)
    summary_sheet["A1"].fill = PatternFill("solid", fgColor=navy)
    summary_sheet["A1"].alignment = Alignment(horizontal="center")
    summary_sheet.row_dimensions[1].height = 30
    summary_sheet["A2"].font = Font(bold=True, color=navy)

    for worksheet in workbook.worksheets:
        if worksheet.title == "Executive Dashboard":
            continue
        header_row = 4 if worksheet.title == "Summary" else 1
        worksheet.freeze_panes = f"A{header_row + 1}"
        worksheet.sheet_view.showGridLines = False
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[header_row]:
            cell.font = Font(bold=True, color=white)
            cell.fill = PatternFill("solid", fgColor=blue)
            cell.alignment = Alignment(vertical="center")
        for column_index in range(1, worksheet.max_column + 1):
            heading = str(worksheet.cell(header_row, column_index).value or "").lower()
            if any(word in heading for word in ["amount", "value", "expected", "paid", "outstanding", "rent collected", "salary", "deduction", "incentive", "payroll"]):
                for row_index in range(header_row + 1, worksheet.max_row + 1):
                    worksheet.cell(row_index, column_index).number_format = '₹#,##0.00;[Red](₹#,##0.00);-'
            elif any(word in heading for word in ["quantity", "payments", "jobs", "services", "complaints", "parts"]):
                for row_index in range(header_row + 1, worksheet.max_row + 1):
                    worksheet.cell(row_index, column_index).number_format = '#,##0;[Red](#,##0);-'
        worksheet.row_dimensions[header_row].height = 24
        for row_number in range(header_row + 1, worksheet.max_row + 1):
            if (row_number - header_row) % 2 == 0:
                for cell in worksheet[row_number]:
                    cell.fill = PatternFill("solid", fgColor=pale_blue)
            for cell in worksheet[row_number]:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        worksheet.auto_filter.ref = (
            f"A{header_row}:{worksheet.cell(header_row, worksheet.max_column).coordinate}"
        )
        worksheet.print_title_rows = f"{header_row}:{header_row}"
        worksheet.page_setup.orientation = "landscape"
        worksheet.page_setup.fitToWidth = 1
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        for column_index, column in enumerate(worksheet.columns, start=1):
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 40)
            worksheet.column_dimensions[get_column_letter(column_index)].width = width

    return workbook


class ReportsSummaryAPIView(APIView):
    permission_classes = [IsAdminOrManager]

    def get(self, request):
        try:
            period = resolve_period(request.query_params)
        except ValueError as exc:
            return Response(
                {"success": False, "message": str(exc)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response({
            "success": True,
            **build_reports(period),
        })


class ReportsExportAPIView(APIView):
    permission_classes = [IsAdminOrManager]

    def get(self, request):
        try:
            period = resolve_period(request.query_params)
        except ValueError as exc:
            return Response(
                {"success": False, "message": str(exc)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        data = build_reports(period)
        workbook = build_workbook(data)
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = (
            f"ari-smart-ro-{period.key}-"
            f"{period.start.isoformat()}-{period.end.isoformat()}.xlsx"
        )
        response = HttpResponse(
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
