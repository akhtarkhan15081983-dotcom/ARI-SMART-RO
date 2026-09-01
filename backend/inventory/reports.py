from io import BytesIO
from secrets import token_hex

import openpyxl
import qrcode
from django.db import transaction
from django.db.models import Count, Sum
from django.http import HttpResponse
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
from rest_framework.response import Response
from rest_framework.views import APIView

from accounts.permissions import IsStaffOperator
from purchase.models import Purchase, PurchaseItem, Supplier

from .models import InventoryAuditLog, InventoryItem, PartRequest


class InventoryCodeGenerationAPIView(APIView):
    permission_classes = [IsStaffOperator]

    @transaction.atomic
    def post(self, request):
        purchase_item_id = request.data.get("purchase_item_id")
        items = list(
            InventoryItem.objects.select_for_update()
            .filter(
                purchase_item_id=purchase_item_id,
                status="PENDING_RECEIPT",
                serial_number__isnull=True,
            )
            .select_related("part")
        )
        requested_count = request.data.get("count")
        if requested_count not in (None, ""):
            try:
                count = int(requested_count)
            except (TypeError, ValueError):
                return Response({"success": False, "message": "Count must be a number."}, status=400)
            if count < 1 or count > len(items):
                return Response({"success": False, "message": "Count exceeds pending quantity."}, status=400)
            items = items[:count]
        if not items:
            return Response({"success": False, "message": "No pending items need QR codes."}, status=409)
        codes = []
        for item in items:
            while True:
                code = f"ARI-{item.part.code}-{token_hex(5).upper()}"
                if not InventoryItem.objects.filter(serial_number=code).exists():
                    break
            item.serial_number = code
            item.barcode = code
            item.save()
            InventoryAuditLog.objects.create(
                inventory_item=item, performed_by=request.user, action="STATUS_CHANGE",
                old_status="PENDING_RECEIPT", new_status="PENDING_RECEIPT",
                serial_number=code, remarks="Internal QR label generated before receipt.",
            )
            codes.append(code)
        return Response({"success": True, "generated": len(codes), "codes": codes})


class InventoryQrLabelsPdfAPIView(APIView):
    permission_classes = [IsStaffOperator]

    def get(self, request):
        queryset = InventoryItem.objects.exclude(serial_number__isnull=True).exclude(
            serial_number=""
        ).select_related("part", "purchase_item__purchase")
        purchase_item_id = request.query_params.get("purchase_item_id")
        if purchase_item_id:
            queryset = queryset.filter(purchase_item_id=purchase_item_id)
        items = list(queryset.order_by("part__code", "id")[:1000])
        if not items:
            return Response({"success": False, "message": "No QR labels are available."}, status=404)
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=A4)
        page_width, page_height = A4
        cols, rows = 3, 8
        card_width, card_height = page_width / cols, page_height / rows
        for index, item in enumerate(items):
            slot = index % (cols * rows)
            if index and slot == 0:
                pdf.showPage()
            col, row = slot % cols, slot // cols
            x, y = col * card_width, page_height - (row + 1) * card_height
            pdf.roundRect(x + 3 * mm, y + 3 * mm, card_width - 6 * mm, card_height - 6 * mm, 2 * mm)
            qr = qrcode.make(item.serial_number)
            image_buffer = BytesIO()
            qr.save(image_buffer, format="PNG")
            image_buffer.seek(0)
            pdf.drawImage(ImageReader(image_buffer), x + 6 * mm, y + 6 * mm, 22 * mm, 22 * mm)
            pdf.setFont("Helvetica-Bold", 8)
            pdf.drawString(x + 30 * mm, y + 25 * mm, item.part.code[:22])
            pdf.setFont("Helvetica", 6.5)
            pdf.drawString(x + 30 * mm, y + 19 * mm, item.part.name[:28])
            pdf.drawString(x + 30 * mm, y + 13 * mm, item.serial_number[:28])
            pdf.drawString(x + 30 * mm, y + 7 * mm, f"INV: {item.purchase_item.purchase.invoice_number}"[:30])
        pdf.save()
        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="ARI_Inventory_QR_Labels.pdf"'
        return response


class InventorySummaryAPIView(APIView):
    permission_classes = [IsStaffOperator]

    def get(self, request):
        by_status = {row["status"]: row["count"] for row in InventoryItem.objects.values("status").annotate(count=Count("id"))}
        return Response({
            "success": True,
            "summary": {
                "total_units": InventoryItem.objects.count(),
                "pending_receipt": by_status.get("PENDING_RECEIPT", 0),
                "in_stock": by_status.get("IN_STOCK", 0),
                "issued": by_status.get("ISSUED", 0),
                "installed": by_status.get("INSTALLED", 0),
                "scrap": by_status.get("SCRAP", 0),
                "pending_requests": PartRequest.objects.filter(status="PENDING").count(),
                "approved_requests": PartRequest.objects.filter(status="APPROVED").count(),
            },
        })


def _sheet(workbook, title, headers, rows):
    sheet = workbook.create_sheet(title)
    sheet.freeze_panes = "A2"
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="075985")
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        sheet.append(list(row))
    sheet.auto_filter.ref = sheet.dimensions
    for column in range(1, len(headers) + 1):
        values = [str(sheet.cell(row=row, column=column).value or "") for row in range(1, min(sheet.max_row, 500) + 1)]
        sheet.column_dimensions[get_column_letter(column)].width = min(max(map(len, values), default=10) + 2, 35)
    return sheet


class InventoryExcelReportAPIView(APIView):
    permission_classes = [IsStaffOperator]

    def get(self, request):
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        summary = InventoryItem.objects.values("status").annotate(units=Count("id")).order_by("status")
        _sheet(workbook, "Executive Summary", ["Metric", "Value"], [
            ("Generated By", request.user.get_full_name() or request.user.phone),
            ("Total Suppliers", Supplier.objects.count()),
            ("Total Purchases", Purchase.objects.count()),
            ("AI-Scanned Invoices", Purchase.objects.filter(entry_source="INVOICE_OCR").count()),
            ("Inventory Units", InventoryItem.objects.count()),
            ("Stock Value", float(PurchaseItem.objects.aggregate(value=Sum("purchase_price"))["value"] or 0)),
            *[(f"Status: {row['status']}", row["units"]) for row in summary],
        ])
        inventory = InventoryItem.objects.select_related("part", "purchase_item__purchase__supplier", "received_by")
        _sheet(workbook, "Stock Ledger", ["ID", "Part Code", "Part", "Serial-QR", "Status", "Invoice", "Supplier", "Received At", "Received By"], [
            (item.id, item.part.code, item.part.name, item.serial_number or "", item.status,
             item.purchase_item.purchase.invoice_number, item.purchase_item.purchase.supplier.name,
             item.received_at.replace(tzinfo=None) if item.received_at else "",
             (item.received_by.get_full_name() or item.received_by.phone) if item.received_by else "")
            for item in inventory
        ])
        purchases = PurchaseItem.objects.select_related("purchase__supplier", "purchase__verified_by", "part")
        _sheet(workbook, "Purchases", ["Invoice", "Date", "Supplier", "Part Code", "Part", "Quantity", "Unit Price", "Total", "Entry Source", "OCR Confidence", "Verified By", "Invoice Image"], [
            (item.purchase.invoice_number, item.purchase.invoice_date, item.purchase.supplier.name,
             item.part.code, item.part.name, item.quantity, float(item.purchase_price), float(item.purchase_price * item.quantity),
             item.purchase.entry_source, float(item.purchase.ocr_confidence),
             (item.purchase.verified_by.get_full_name() or item.purchase.verified_by.phone) if item.purchase.verified_by else "",
             item.purchase.invoice_image.name if item.purchase.invoice_image else "")
            for item in purchases
        ])
        requests = PartRequest.objects.select_related("engineer__user", "part", "reviewed_by")
        _sheet(workbook, "Part Requests", ["Request", "Employee", "Part Code", "Part", "Qty", "Status", "Requested", "Reviewed By", "Review Note"], [
            (row.id, row.engineer.user.get_full_name(), row.part.code, row.part.name, row.quantity,
             row.status, row.created_at.replace(tzinfo=None),
             (row.reviewed_by.get_full_name() or row.reviewed_by.phone) if row.reviewed_by else "", row.review_remarks)
            for row in requests
        ])
        audit = InventoryAuditLog.objects.select_related("inventory_item__part", "engineer__user", "performed_by")
        _sheet(workbook, "Audit Trail", ["Time", "Action", "Part", "Serial", "Old Status", "New Status", "Engineer", "Performed By", "Remarks"], [
            (row.created_at.replace(tzinfo=None), row.action, row.inventory_item.part.name, row.serial_number,
             row.old_status, row.new_status, row.engineer.user.get_full_name() if row.engineer else "",
             (row.performed_by.get_full_name() or row.performed_by.phone) if row.performed_by else "", row.remarks)
            for row in audit
        ])
        output = BytesIO()
        workbook.save(output)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="ARI_Professional_Inventory_Report.xlsx"'
        return response
