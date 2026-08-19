import os
from collections import Counter, defaultdict

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from customers.models import Customer
from employees.models import EmployeeProfile


class Command(BaseCommand):
    help = "Sync customer assignments from Excel. Dry-run by default."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True)
        parser.add_argument(
            "--apply",
            action="store_true",
            help="Actually update customer assignments.",
        )

    @staticmethod
    def norm(value):
        if value is None:
            return ""
        return " ".join(str(value).strip().upper().split())

    def handle(self, *args, **options):
        file_path = options["file"]
        apply_changes = options["apply"]

        if not os.path.exists(file_path):
            raise CommandError(f"Excel file not found: {file_path}")

        # ---------------------------------------------------------
        # EXCEL EMPLOYEE NAME -> DATABASE EMPLOYEE FULL NAME
        # ---------------------------------------------------------
        employee_aliases = {
            "HUSNAIN GAZI": "HUSNAIN GAZI",
            "HASNAIN": "HUSNAIN GAZI",
            "HUSNAIN": "HUSNAIN GAZI",

            "RAJKUMAR KUSHWAH": "RAJKUMAR KUSHWAH",
            "RAJKUMAR": "RAJKUMAR KUSHWAH",

            "RUPESH BHAGHEL": "RUPESH BHAGHEL",
            "RUPESH": "RUPESH BHAGHEL",

            "RAMA AASRE": "RAMA AASRE",
            "RAMA": "RAMA AASRE",

            "FAIZAN": "FAIZAN",
            "FAIZAN KHAN": "FAIZAN",
        }

        # ---------------------------------------------------------
        # LOAD ACTIVE EMPLOYEE PROFILES
        #
        # IMPORTANT:
        # We intentionally DO NOT filter designation="ENGINEER"
        # because FAIZAN is OFFICE staff.
        # ---------------------------------------------------------
        profiles = {}

        for profile in EmployeeProfile.objects.select_related("user").filter(
            is_active=True
        ):
            full_name = self.norm(profile.user.get_full_name())

            if full_name:
                profiles[full_name] = profile

        # ---------------------------------------------------------
        # BUILD FINAL EMPLOYEE MAP
        # ---------------------------------------------------------
        employee_map = {}

        for excel_name, db_name in employee_aliases.items():
            profile = profiles.get(db_name)

            if profile is not None:
                employee_map[excel_name] = profile

        # ---------------------------------------------------------
        # REQUIRED EMPLOYEES
        # ---------------------------------------------------------
        required = [
            "HUSNAIN GAZI",
            "RAJKUMAR KUSHWAH",
            "RUPESH BHAGHEL",
            "RAMA AASRE",
            "FAIZAN",
        ]

        missing = [
            name for name in required
            if name not in profiles
        ]

        if missing:
            raise CommandError(
                "Required employee profile missing: "
                + ", ".join(missing)
            )

        # ---------------------------------------------------------
        # PRINT EMPLOYEES FOUND
        # ---------------------------------------------------------
        self.stdout.write("")
        self.stdout.write("=" * 70)
        self.stdout.write("EMPLOYEES FOUND")
        self.stdout.write("=" * 70)

        for name in required:
            profile = profiles[name]

            self.stdout.write(
                f"{name} -> "
                f"Profile ID: {profile.id} | "
                f"Employee ID: {profile.employee_id} | "
                f"Role: {profile.user.role} | "
                f"Designation: {profile.designation}"
            )

        # ---------------------------------------------------------
        # LOAD EXCEL
        # ---------------------------------------------------------
        workbook = load_workbook(
            filename=file_path,
            read_only=True,
            data_only=True,
        )

        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)

        try:
            next(rows)
            headers = list(next(rows))
        except StopIteration:
            workbook.close()
            raise CommandError(
                "Excel does not contain the expected header row."
            )

        headers = [
            str(value).strip() if value is not None else ""
            for value in headers
        ]

        def col(name):
            wanted = name.lower().strip()

            for index, header in enumerate(headers):
                if header.lower().strip() == wanted:
                    return index

            return None

        old_i = col("card number")
        name_i = col("CUSTUMER NAME")
        emp_i = col("employee")

        if old_i is None:
            workbook.close()
            raise CommandError(
                "Excel column 'card number' not found."
            )

        if name_i is None:
            workbook.close()
            raise CommandError(
                "Excel column 'CUSTUMER NAME' not found."
            )

        if emp_i is None:
            workbook.close()
            raise CommandError(
                "Excel column 'employee' not found."
            )

        # ---------------------------------------------------------
        # DATABASE CUSTOMERS BY OLD CARD NUMBER
        # ---------------------------------------------------------
        db = {}
        duplicate_db = defaultdict(list)

        for customer in Customer.objects.exclude(
            old_card_number=""
        ):
            key = self.norm(customer.old_card_number)

            if key in db:
                duplicate_db[key].append(customer.id)
            else:
                db[key] = customer

        # ---------------------------------------------------------
        # COUNTERS
        # ---------------------------------------------------------
        matched = Counter()
        changed = Counter()
        unchanged = Counter()
        unmatched = Counter()

        changes = []
        skipped = []

        # ---------------------------------------------------------
        # READ EXCEL
        # ---------------------------------------------------------
        for excel_row, row in enumerate(rows, start=3):

            employee_raw = (
                row[emp_i]
                if emp_i < len(row)
                else None
            )

            employee_excel = self.norm(employee_raw)

            # Ignore employees that are not part of our assignment
            # system, e.g. ONLINE, KULDEEP, KAINNETH, etc.
            if employee_excel not in employee_map:
                continue

            old_card = (
                self.norm(row[old_i])
                if old_i < len(row)
                else ""
            )

            excel_name = (
                str(row[name_i]).strip()
                if name_i < len(row)
                and row[name_i] is not None
                else ""
            )

            # Missing card number
            if not old_card:
                unmatched[employee_excel] += 1

                skipped.append(
                    (
                        excel_row,
                        employee_excel,
                        old_card,
                        excel_name,
                        "missing card number",
                    )
                )

                continue

            # Duplicate card in database
            if old_card in duplicate_db:

                unmatched[employee_excel] += 1

                skipped.append(
                    (
                        excel_row,
                        employee_excel,
                        old_card,
                        excel_name,
                        "duplicate card number in database",
                    )
                )

                continue

            # Customer not found
            customer = db.get(old_card)

            if customer is None:

                unmatched[employee_excel] += 1

                skipped.append(
                    (
                        excel_row,
                        employee_excel,
                        old_card,
                        excel_name,
                        "customer not found in database",
                    )
                )

                continue

            # Employee profile
            profile = employee_map[employee_excel]

            matched[employee_excel] += 1

            # Already assigned correctly
            if customer.assigned_engineer_id == profile.id:

                unchanged[employee_excel] += 1

            # Assignment needs to change
            else:

                changed[employee_excel] += 1

                changes.append(
                    (
                        excel_row,
                        old_card,
                        customer,
                        employee_excel,
                        profile,
                    )
                )

        workbook.close()

        # ---------------------------------------------------------
        # TOTALS
        # ---------------------------------------------------------
        total_matched = sum(matched.values())
        total_changed = sum(changed.values())
        total_unchanged = sum(unchanged.values())
        total_unmatched = sum(unmatched.values())

        # ---------------------------------------------------------
        # REPORT
        # ---------------------------------------------------------
        self.stdout.write("")
        self.stdout.write("=" * 70)

        self.stdout.write(
            "CUSTOMER ASSIGNMENT "
            + ("APPLY" if apply_changes else "DRY RUN")
        )

        self.stdout.write("=" * 70)

        for employee in required:

            self.stdout.write(
                f"{employee} | "
                f"Matched: {matched[employee]} | "
                f"Will Change: {changed[employee]} | "
                f"Already Correct: {unchanged[employee]} | "
                f"Unmatched: {unmatched[employee]}"
            )

        self.stdout.write("")
        self.stdout.write(
            f"TOTAL MATCHED  : {total_matched}"
        )

        self.stdout.write(
            f"TOTAL TO CHANGE: {total_changed}"
        )

        self.stdout.write(
            f"ALREADY CORRECT: {total_unchanged}"
        )

        self.stdout.write(
            f"UNMATCHED      : {total_unmatched}"
        )

        # ---------------------------------------------------------
        # SHOW CHANGES
        # ---------------------------------------------------------
        if changes:

            self.stdout.write("")
            self.stdout.write("=" * 70)
            self.stdout.write("FIRST CHANGES")
            self.stdout.write("=" * 70)

            for (
                excel_row,
                old_card,
                customer,
                employee,
                profile,
            ) in changes[:50]:

                old_profile = customer.assigned_engineer_id

                self.stdout.write(
                    f"Row {excel_row} | "
                    f"Card {old_card} | "
                    f"{customer.name} | "
                    f"{employee} | "
                    f"OLD Profile: {old_profile} -> "
                    f"NEW Profile: {profile.id}"
                )

        # ---------------------------------------------------------
        # SHOW SKIPPED
        # ---------------------------------------------------------
        if skipped:

            self.stdout.write("")
            self.stdout.write("=" * 70)
            self.stdout.write("FIRST SKIPPED")
            self.stdout.write("=" * 70)

            for (
                excel_row,
                employee,
                old_card,
                name,
                reason,
            ) in skipped[:50]:

                self.stdout.write(
                    f"Row {excel_row} | "
                    f"{employee} | "
                    f"Card {old_card or '(empty)'} | "
                    f"{name or '(empty)'} | "
                    f"{reason}"
                )

        # ---------------------------------------------------------
        # DRY RUN
        # ---------------------------------------------------------
        if not apply_changes:

            self.stdout.write("")
            self.stdout.write("=" * 70)
            self.stdout.write("DRY RUN COMPLETE")
            self.stdout.write("=" * 70)

            self.stdout.write(
                "NO DATABASE RECORDS WERE CHANGED."
            )

            self.stdout.write(
                "Use --apply after reviewing the preview."
            )

            return

        # ---------------------------------------------------------
        # APPLY CHANGES
        # ---------------------------------------------------------
        with transaction.atomic():

            for (
                _,
                _,
                customer,
                _,
                profile,
            ) in changes:

                customer.assigned_engineer = profile

                customer.save(
                    update_fields=["assigned_engineer"]
                )

        # ---------------------------------------------------------
        # FINAL
        # ---------------------------------------------------------
        self.stdout.write("")
        self.stdout.write("=" * 70)
        self.stdout.write("ASSIGNMENT SYNC COMPLETE")
        self.stdout.write("=" * 70)

        self.stdout.write(
            f"Changed           : {len(changes)}"
        )

        self.stdout.write(
            f"Already correct   : {total_unchanged}"
        )

        self.stdout.write(
            f"Skipped/unmatched : {total_unmatched}"
        )

        self.stdout.write(
            self.style.SUCCESS(
                "Customer assignments synchronized successfully."
            )
        )