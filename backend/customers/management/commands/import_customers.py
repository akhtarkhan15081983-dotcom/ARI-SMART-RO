import os
from datetime import datetime, date
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from customers.models import Customer, CustomerRentHistory


class Command(BaseCommand):

    help = "Import ARI SMART RO customers from Excel"

    def add_arguments(self, parser):

        parser.add_argument(
            "--file",
            required=True,
            help="Path to the customer Excel file",
        )

        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Check Excel without saving anything",
        )

        parser.add_argument(
            "--limit",
            type=int,
            default=0,
            help="Import first N customers. 0 = all",
        )

    def handle(self, *args, **options):

        file_path = options["file"]
        dry_run = options["dry_run"]
        limit = options["limit"]

        if not os.path.exists(file_path):
            raise CommandError(
                f"Excel file not found: {file_path}"
            )

        self.stdout.write(
            self.style.SUCCESS(
                f"Excel file found: {file_path}"
            )
        )

        workbook = load_workbook(
            filename=file_path,
            read_only=True,
            data_only=True,
        )

        worksheet = workbook[workbook.sheetnames[0]]

        self.stdout.write(
            self.style.SUCCESS(
                f"Using sheet: {worksheet.title}"
            )
        )

        rows = worksheet.iter_rows(values_only=True)

        # -------------------------------------------------
        # ROW 1 = SUMMARY
        # ROW 2 = ACTUAL HEADER
        # -------------------------------------------------

        try:
            next(rows)
            headers = list(next(rows))
        except StopIteration:

            workbook.close()

            raise CommandError(
                "Excel does not contain a valid header row."
            )

        # Keep original values.
        # Do NOT convert datetime headers to strings.
        normalized_headers = []

        for value in headers:

            if isinstance(value, str):
                normalized_headers.append(
                    value.strip()
                )

            else:
                normalized_headers.append(value)

        headers = normalized_headers

        # -------------------------------------------------
        # FIND COLUMN
        # -------------------------------------------------

        def find_column(name):

            target = name.lower().strip()

            for index, header in enumerate(headers):

                if isinstance(header, str):

                    if header.lower().strip() == target:
                        return index

            return None

        sno_col = find_column("S.No.")
        old_card_col = find_column("card number")
        new_card_col = find_column("new card no.")
        name_col = find_column("CUSTUMER NAME")
        phone_col = find_column("Contect No.")
        area_col = find_column("AREA")
        address_col = find_column("adress")
        installation_date_col = find_column(
            "DATE OF INSTALLATION"
        )
        installation_charge_col = find_column(
            "Instalation Charge"
        )
        rent_col = find_column("Rent")
        employee_col = find_column("employee")
        installer_col = find_column("installer")

        required_columns = {
            "S.No.": sno_col,
            "card number": old_card_col,
            "CUSTUMER NAME": name_col,
            "Contect No.": phone_col,
            "AREA": area_col,
            "adress": address_col,
            "DATE OF INSTALLATION": installation_date_col,
            "Instalation Charge": installation_charge_col,
            "Rent": rent_col,
        }

        missing = [
            name
            for name, index in required_columns.items()
            if index is None
        ]

        if missing:

            workbook.close()

            raise CommandError(
                "Missing required Excel columns: "
                + ", ".join(missing)
            )

        # -------------------------------------------------
        # PAYMENT COLUMNS
        #
        # Only REAL datetime headers are payment months.
        #
        # DEC / FEB / blank headers are ignored.
        # -------------------------------------------------

        payment_columns = []

        for index in range(
            rent_col + 1,
            len(headers)
        ):

            header = headers[index]

            payment_month = None

            if isinstance(header, datetime):

                payment_month = header.date()

            elif isinstance(header, date):

                payment_month = header

            if payment_month:

                payment_columns.append(
                    (
                        index,
                        payment_month,
                        header,
                    )
                )

        self.stdout.write("")

        self.stdout.write(
            self.style.SUCCESS(
                f"Valid payment months found: "
                f"{len(payment_columns)}"
            )
        )

        for index, payment_month, original_header in payment_columns:

            self.stdout.write(
                f"Column {index + 1}: "
                f"{payment_month}"
            )

        # -------------------------------------------------
        # HELPERS
        # -------------------------------------------------

        def clean_string(value):

            if value is None:
                return ""

            return str(value).strip()

        def clean_phone(value):

            if value is None:
                return ""

            if isinstance(value, float):

                value = int(value)

            value = str(value).strip()

            if value.endswith(".0"):

                value = value[:-2]

            value = "".join(
                character
                for character in value
                if character.isdigit()
            )

            return value[-10:]

        def decimal_value(value):

            if value is None:
                return Decimal("0")

            if isinstance(value, str):

                value = value.strip()

                if not value:
                    return Decimal("0")

            try:

                return Decimal(str(value))

            except (
                InvalidOperation,
                ValueError,
            ):

                return Decimal("0")

        def parse_date(value):

            if value is None:
                return None

            if isinstance(value, datetime):

                return value.date()

            if isinstance(value, date):

                return value

            if isinstance(value, str):

                value = value.strip()

                formats = [
                    "%d/%m/%Y",
                    "%d-%m-%Y",
                    "%Y-%m-%d",
                    "%d/%m/%y",
                ]

                for fmt in formats:

                    try:

                        return datetime.strptime(
                            value,
                            fmt,
                        ).date()

                    except ValueError:

                        continue

            return None

        # -------------------------------------------------
        # COUNTERS
        # -------------------------------------------------

        processed = 0
        valid = 0
        skipped = 0
        duplicate = 0
        imported_history = 0

        preview = []

        # -------------------------------------------------
        # CUSTOMER LOOP
        # -------------------------------------------------

        for excel_row_number, row in enumerate(
            rows,
            start=3,
        ):

            if limit and processed >= limit:
                break

            processed += 1

            row = list(row)

            # Ensure row length
            if len(row) < len(headers):

                row.extend(
                    [None] *
                    (
                        len(headers) - len(row)
                    )
                )

            # -------------------------------------------------
            # CUSTOMER DATA
            # -------------------------------------------------

            name = clean_string(
                row[name_col]
            )

            phone = clean_phone(
                row[phone_col]
            )

            old_card = clean_string(
                row[old_card_col]
            )

            new_card = ""

            if new_card_col is not None:

                new_card = clean_string(
                    row[new_card_col]
                )

            area = clean_string(
                row[area_col]
            )

            address = clean_string(
                row[address_col]
            )

            installation_date = parse_date(
                row[installation_date_col]
            )

            installation_charge = decimal_value(
                row[installation_charge_col]
            )

            monthly_rent = decimal_value(
                row[rent_col]
            )

            # -------------------------------------------------
            # VALIDATION
            # -------------------------------------------------

            if not name:

                skipped += 1

                self.stdout.write(
                    self.style.WARNING(
                        f"Row {excel_row_number}: "
                        f"name missing"
                    )
                )

                continue

            if not phone:

                skipped += 1

                self.stdout.write(
                    self.style.WARNING(
                        f"Row {excel_row_number}: "
                        f"{name} - phone missing"
                    )
                )

                continue

            valid += 1

            # -------------------------------------------------
            # FIND EXISTING CUSTOMER
            # -------------------------------------------------

            existing = None

            if old_card:

                existing = (
                    Customer.objects
                    .filter(
                        old_card_number=old_card
                    )
                    .first()
                )

            if existing is None and new_card:

                existing = (
                    Customer.objects
                    .filter(
                        card_number=new_card
                    )
                    .first()
                )

            if existing is None:

                existing = (
                    Customer.objects
                    .filter(
                        phone=phone
                    )
                    .first()
                )

            if existing:

                duplicate += 1

            # -------------------------------------------------
            # DRY RUN PREVIEW
            # -------------------------------------------------

            if len(preview) < 5:

                preview.append(
                    (
                        excel_row_number,
                        name,
                        phone,
                        old_card,
                        new_card,
                        monthly_rent,
                        installation_date,
                    )
                )

            # -------------------------------------------------
            # DATABASE
            # -------------------------------------------------

            if not dry_run:

                with transaction.atomic():

                    # =========================================
                    # EXISTING CUSTOMER
                    # =========================================

                    if existing:

                        customer = existing

                        customer.name = name
                        customer.phone = phone
                        customer.area = area
                        customer.address = address

                        customer.installation_date = (
                            installation_date
                        )

                        customer.installation_charge = (
                            installation_charge
                        )

                        customer.monthly_rent = (
                            monthly_rent
                        )

                        if old_card:

                            customer.old_card_number = (
                                old_card
                            )

                        if new_card:

                            customer.card_number = (
                                new_card
                            )

                        if not customer.city:

                            customer.city = "Agra"

                        if not customer.state:

                            customer.state = (
                                "Uttar Pradesh"
                            )

                        if not customer.pincode:

                            customer.pincode = "282001"

                        if not customer.ro_model:

                            customer.ro_model = "UNKNOWN"

                        customer.save()

                    # =========================================
                    # NEW CUSTOMER
                    # =========================================

                    else:

                        customer = Customer(
                            name=name,
                            phone=phone,
                            old_card_number=old_card,
                            card_number=new_card,
                            area=area,
                            address=address,
                            city="Agra",
                            state="Uttar Pradesh",
                            pincode="282001",
                            ro_model="UNKNOWN",
                            installation_charge=(
                                installation_charge
                            ),
                            monthly_rent=(
                                monthly_rent
                            ),
                            security_deposit=Decimal("0"),
                            installation_date=(
                                installation_date
                            ),
                            is_active=True,
                        )

                        customer.save()

                        # Customer.save() generates
                        # card_number automatically
                        # when new_card is empty.
                        #
                        # If Excel has a new card number,
                        # preserve it.

                        if new_card:

                            customer.card_number = (
                                new_card
                            )

                            customer.save(
                                update_fields=[
                                    "card_number"
                                ]
                            )

                    # =========================================
                    # RENT HISTORY
                    # =========================================

                    customer_history_count = 0

                    for (
                        column_index,
                        payment_month,
                        original_header,
                    ) in payment_columns:

                        value = row[column_index]

                        # Empty cell = no payment record
                        if value is None:
                            continue

                        raw_value = clean_string(
                            value
                        )

                        paid_amount = decimal_value(
                            value
                        )

                        CustomerRentHistory.objects.update_or_create(

                            customer=customer,

                            rent_month=payment_month,

                            defaults={
                                "expected_rent": monthly_rent,
                                "paid_amount": paid_amount,
                                "raw_value": raw_value,
                                "remarks": (
                                    "Imported from Excel"
                                ),
                            },
                        )

                        customer_history_count += 1
                        imported_history += 1

                    # -------------------------------------------------
                    # DEBUG FOR FIRST 3
                    # -------------------------------------------------

                    if processed <= 3:

                        self.stdout.write(
                            self.style.SUCCESS(
                                f"Row {excel_row_number}: "
                                f"{name} -> "
                                f"{customer_history_count} "
                                f"rent history records"
                            )
                        )

            # -------------------------------------------------
            # PROGRESS
            # -------------------------------------------------

            if processed % 100 == 0:

                self.stdout.write(
                    f"Processed: {processed}"
                )

        workbook.close()

        # -------------------------------------------------
        # SUMMARY
        # -------------------------------------------------

        self.stdout.write("")

        self.stdout.write(
            self.style.SUCCESS(
                "========== IMPORT SUMMARY =========="
            )
        )

        self.stdout.write(
            f"Rows processed       : {processed}"
        )

        self.stdout.write(
            f"Valid customers      : {valid}"
        )

        self.stdout.write(
            f"Skipped              : {skipped}"
        )

        self.stdout.write(
            f"Existing/duplicate   : {duplicate}"
        )

        self.stdout.write(
            f"Rent history records : {imported_history}"
        )

        self.stdout.write("")

        self.stdout.write(
            self.style.SUCCESS(
                "========== PREVIEW =========="
            )
        )

        for item in preview:

            (
                excel_row,
                name,
                phone,
                old_card,
                new_card,
                rent,
                install_date,
            ) = item

            self.stdout.write(
                f"Row {excel_row} | "
                f"{name} | "
                f"Phone: {phone} | "
                f"Old: {old_card} | "
                f"New: {new_card} | "
                f"Rent: ₹{rent} | "
                f"Install: {install_date}"
            )

        if dry_run:

            self.stdout.write("")

            self.stdout.write(
                self.style.WARNING(
                    "DRY RUN COMPLETE"
                )
            )

            self.stdout.write(
                self.style.WARNING(
                    "NO DATABASE RECORDS WERE CHANGED."
                )
            )

        else:

            self.stdout.write("")

            self.stdout.write(
                self.style.SUCCESS(
                    "CUSTOMER IMPORT COMPLETE"
                )
            )