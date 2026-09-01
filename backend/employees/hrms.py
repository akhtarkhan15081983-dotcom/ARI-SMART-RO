import calendar
from io import BytesIO
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP

from django.db import transaction
from django.db.models import Sum
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from attendance.models import Attendance
from installation.models import Installation
from .models import EmployeeProfile, Holiday, HRPolicy, LeaveRequest, PayrollRecord


MONEY = Decimal("0.01")


def _money(value):
    return Decimal(value).quantize(MONEY, rounding=ROUND_HALF_UP)


def _role(user, *roles):
    return getattr(user, "role", "") in roles


def _month(value):
    parsed = datetime.strptime(value, "%Y-%m").date()
    return parsed.replace(day=1)


def calculate_payroll(employee, payroll_month):
    policy = HRPolicy.current()
    days_in_month = calendar.monthrange(payroll_month.year, payroll_month.month)[1]
    month_end = payroll_month.replace(day=days_in_month)
    start = max(payroll_month, employee.joining_date)
    daily_rate = Decimal(employee.salary) / Decimal(days_in_month)
    eligible_days = Decimal((month_end - start).days + 1)
    payable_base = _money(daily_rate * eligible_days)
    hourly_rate = daily_rate / Decimal(policy.daily_work_hours)
    attendance = list(Attendance.objects.filter(employee=employee, date__range=(start, month_end)))
    attendance_by_date = {row.date: row for row in attendance}
    holiday_dates = set(Holiday.objects.filter(date__range=(start, month_end)).values_list("date", flat=True))
    leaves = list(LeaveRequest.objects.filter(employee=employee, status="APPROVED", start_date__lte=month_end, end_date__gte=start).order_by("created_at"))

    full_paid_remaining = policy.monthly_paid_leaves
    half_paid_remaining = policy.monthly_paid_half_days
    paid_leave_dates, unpaid_leave_units = set(), Decimal("0")
    for leave in leaves:
        current = max(start, leave.start_date)
        final = min(month_end, leave.end_date)
        while current <= final:
            if leave.leave_type == "HALF_DAY":
                if half_paid_remaining > 0:
                    paid_leave_dates.add((current, "HALF"))
                    half_paid_remaining -= 1
                else:
                    unpaid_leave_units += Decimal("0.5")
            elif full_paid_remaining > 0:
                paid_leave_dates.add((current, "FULL"))
                full_paid_remaining -= 1
            else:
                unpaid_leave_units += Decimal("1")
            current += timedelta(days=1)

    late_days = half_day_units = 0
    overtime_hours = Decimal("0")
    absent_days = 0
    current = start
    while current <= month_end:
        row = attendance_by_date.get(current)
        if current in holiday_dates and row is None:
            current += timedelta(days=1)
            continue
        full_leave = (current, "FULL") in paid_leave_dates
        half_leave = (current, "HALF") in paid_leave_dates
        if row is None:
            if not full_leave:
                absent_days += Decimal("0.5") if half_leave else Decimal("1")
        else:
            if row.check_in:
                local_checkin = timezone.localtime(row.check_in).time().replace(tzinfo=None)
                if local_checkin >= policy.half_day_cutoff:
                    half_day_units += 1
                elif local_checkin > policy.office_start_time:
                    late_days += 1
            if row.status == "HALF_DAY" and (not row.check_in or timezone.localtime(row.check_in).time().replace(tzinfo=None) < policy.half_day_cutoff):
                half_day_units += 1
            overtime_hours += max(Decimal("0"), Decimal(row.working_hours) - Decimal(policy.daily_work_hours))
        current += timedelta(days=1)

    absence_deduction = _money((absent_days + unpaid_leave_units) * daily_rate)
    half_day_deduction = _money(Decimal(half_day_units) * daily_rate / Decimal("2"))
    late_penalty = _money(Decimal(late_days) * policy.late_penalty_amount)
    overtime_amount = _money(overtime_hours * hourly_rate)

    completed = Installation.objects.filter(engineer=employee, status="COMPLETED", completed_date__isnull=False)
    rent_count = sale_count = 0
    for installation in completed.only("business_type", "completed_date"):
        completed_local = timezone.localtime(installation.completed_date).date()
        distance = (payroll_month.year - completed_local.year) * 12 + payroll_month.month - completed_local.month
        if installation.business_type == "RENT" and 1 <= distance <= policy.rent_installation_incentive_months:
            rent_count += 1
        elif installation.business_type == "SALE" and distance == 1:
            sale_count += 1
    rent_incentive = _money(Decimal(rent_count) * policy.rent_installation_monthly_incentive)
    sale_incentive = _money(Decimal(sale_count) * policy.sale_installation_incentive)
    net = _money(payable_base - late_penalty - half_day_deduction - absence_deduction + overtime_amount + rent_incentive + sale_incentive)
    return {
        "base_salary": _money(employee.salary), "payable_base": payable_base,
        "late_days": late_days, "late_penalty": late_penalty,
        "half_day_deduction": half_day_deduction, "absence_deduction": absence_deduction,
        "overtime_hours": overtime_hours.quantize(MONEY), "overtime_amount": overtime_amount,
        "rent_incentive": rent_incentive, "sale_incentive": sale_incentive, "net_salary": max(Decimal("0"), net),
        "snapshot": {"calendar_days": days_in_month, "absent_days": str(absent_days), "unpaid_leave_units": str(unpaid_leave_units), "rent_installations": rent_count, "sale_installations": sale_count, "daily_rate": str(_money(daily_rate)), "hourly_rate": str(_money(hourly_rate))},
    }


class EmployeeHrmsDashboardAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            employee = request.user.employee_profile
        except (AttributeError, EmployeeProfile.DoesNotExist):
            return Response({"detail": "Employee profile not found."}, status=404)

        today = timezone.localdate()
        month_start = today.replace(day=1)
        policy = HRPolicy.current()
        attendance = Attendance.objects.filter(
            employee=employee, date__range=(month_start, today)
        )
        rows = list(attendance)
        total_hours = sum((Decimal(row.working_hours) for row in rows), Decimal("0"))
        overtime_hours = sum(
            (max(Decimal("0"), Decimal(row.working_hours) - Decimal(policy.daily_work_hours)) for row in rows),
            Decimal("0"),
        )
        late_days = 0
        half_days = 0
        for row in rows:
            if row.status == "HALF_DAY":
                half_days += 1
            if row.check_in:
                check_in = timezone.localtime(row.check_in).time().replace(tzinfo=None)
                if check_in >= policy.half_day_cutoff:
                    if row.status != "HALF_DAY":
                        half_days += 1
                elif check_in > policy.office_start_time:
                    late_days += 1

        month_leaves = LeaveRequest.objects.filter(
            employee=employee, start_date__lte=today.replace(day=calendar.monthrange(today.year, today.month)[1]),
            end_date__gte=month_start,
        )
        approved = month_leaves.filter(status="APPROVED")
        full_used = sum(
            ((min(row.end_date, today) - max(row.start_date, month_start)).days + 1 for row in approved if row.leave_type == "FULL_DAY"),
            0,
        )
        half_used = sum(
            ((min(row.end_date, today) - max(row.start_date, month_start)).days + 1 for row in approved if row.leave_type == "HALF_DAY"),
            0,
        )
        current_installations = Installation.objects.filter(
            engineer=employee,
            status="COMPLETED",
            completed_date__date__range=(month_start, today),
        )
        new_rent = current_installations.filter(business_type="RENT").count()
        new_sales = current_installations.filter(business_type="SALE").count()
        latest_payroll = PayrollRecord.objects.filter(employee=employee).first()

        return Response({
            "employee": {
                "name": employee.user.get_full_name() or employee.user.phone,
                "employee_id": employee.employee_id,
                "designation": employee.get_designation_display(),
                "joining_date": employee.joining_date,
            },
            "month": month_start.strftime("%Y-%m"),
            "attendance": {
                "present_days": attendance.exclude(status="ABSENT").count(),
                "half_days": half_days,
                "late_days": late_days,
                "total_hours": str(total_hours.quantize(MONEY)),
                "overtime_hours": str(overtime_hours.quantize(MONEY)),
                "pending_selfie_reviews": attendance.filter(identity_review_status="PENDING").count(),
            },
            "leave_balance": {
                "paid_full_allowed": policy.monthly_paid_leaves,
                "paid_full_used": full_used,
                "paid_full_remaining": max(0, policy.monthly_paid_leaves - full_used),
                "paid_half_allowed": policy.monthly_paid_half_days,
                "paid_half_used": half_used,
                "paid_half_remaining": max(0, policy.monthly_paid_half_days - half_used),
                "pending_requests": month_leaves.filter(status="PENDING").count(),
            },
            "earnings": {
                "monthly_salary": str(_money(employee.salary)),
                "new_rent_installations": new_rent,
                "new_sales": new_sales,
                "future_rent_monthly_incentive": str(_money(Decimal(new_rent) * policy.rent_installation_monthly_incentive)),
                "next_salary_sale_incentive": str(_money(Decimal(new_sales) * policy.sale_installation_incentive)),
            },
            "policy": {
                "office_start_time": policy.office_start_time.strftime("%I:%M %p"),
                "daily_work_hours": str(policy.daily_work_hours),
                "late_penalty": str(policy.late_penalty_amount),
                "leave_notice_days": policy.leave_notice_days,
            },
            "latest_payroll": None if latest_payroll is None else {
                "month": latest_payroll.payroll_month,
                "net_salary": str(latest_payroll.net_salary),
                "status": latest_payroll.status,
                "paid_at": latest_payroll.paid_at,
            },
        })


class HolidayAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        year = request.query_params.get("year")
        rows = Holiday.objects.select_related("declared_by")
        if year:
            try:
                rows = rows.filter(date__year=int(year))
            except ValueError:
                return Response({"detail": "Year must be numeric."}, status=400)
        return Response({"holidays": [{
            "id": row.id,
            "date": row.date,
            "name": row.name,
            "description": row.description,
            "is_paid": row.is_paid,
            "declared_by": row.declared_by.get_full_name() or row.declared_by.phone,
        } for row in rows[:500]]})

    def post(self, request):
        if not _role(request.user, "ADMIN", "OFFICE"):
            return Response({"detail": "Only admin or office can declare a holiday."}, status=403)
        try:
            holiday_date = date.fromisoformat(request.data.get("date", ""))
        except ValueError:
            return Response({"detail": "Valid holiday date is required."}, status=400)
        name = (request.data.get("name") or "").strip()
        if not name:
            return Response({"detail": "Holiday name is required."}, status=400)
        row, created = Holiday.objects.update_or_create(
            date=holiday_date,
            defaults={
                "name": name[:120],
                "description": (request.data.get("description") or "").strip()[:300],
                "is_paid": bool(request.data.get("is_paid", True)),
                "declared_by": request.user,
            },
        )
        return Response({"id": row.id, "detail": "Holiday declared successfully."}, status=201 if created else 200)


class HolidayDetailAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, holiday_id):
        if not _role(request.user, "ADMIN", "OFFICE"):
            return Response({"detail": "Only admin or office can remove a holiday."}, status=403)
        deleted, _ = Holiday.objects.filter(pk=holiday_id).delete()
        if not deleted:
            return Response({"detail": "Holiday not found."}, status=404)
        return Response(status=204)


class LeaveRequestAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        rows = LeaveRequest.objects.select_related("employee__user")
        if not _role(request.user, "ADMIN", "MANAGER", "OFFICE"):
            rows = rows.filter(employee__user=request.user)
        return Response({"leaves": [{"id": r.id, "employee": r.employee.employee_id, "employee_name": r.employee.user.get_full_name(), "type": r.leave_type, "start_date": r.start_date, "end_date": r.end_date, "reason": r.reason, "status": r.status, "is_paid": r.is_paid, "review_note": r.review_note} for r in rows[:500]]})

    def post(self, request):
        try:
            employee = request.user.employee_profile
            start = date.fromisoformat(request.data.get("start_date", ""))
            end = date.fromisoformat(request.data.get("end_date", ""))
        except (AttributeError, ValueError):
            return Response({"detail": "Valid employee and leave dates are required."}, status=400)
        policy = HRPolicy.current()
        if start < timezone.localdate() + timedelta(days=policy.leave_notice_days):
            return Response({"detail": "Leave must be requested at least 1 day in advance."}, status=400)
        if end < start:
            return Response({"detail": "End date cannot be before start date."}, status=400)
        if LeaveRequest.objects.filter(employee=employee, status__in=["PENDING", "APPROVED"], start_date__lte=end, end_date__gte=start).exists():
            return Response({"detail": "A leave request already exists for these dates."}, status=400)
        row = LeaveRequest.objects.create(employee=employee, leave_type=request.data.get("leave_type", "FULL_DAY"), start_date=start, end_date=end, reason=(request.data.get("reason") or "").strip())
        return Response({"id": row.id, "status": row.status, "detail": "Leave request submitted for approval."}, status=201)


class LeaveReviewAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, leave_id):
        if not _role(request.user, "ADMIN", "MANAGER", "OFFICE"):
            return Response({"detail": "Not permitted."}, status=403)
        row = LeaveRequest.objects.filter(pk=leave_id, status="PENDING").first()
        if row is None:
            return Response({"detail": "Pending leave not found."}, status=404)
        decision = request.data.get("status")
        if decision not in ("APPROVED", "REJECTED"):
            return Response({"detail": "Status must be APPROVED or REJECTED."}, status=400)
        row.status, row.reviewed_by, row.reviewed_at = decision, request.user, timezone.now()
        row.review_note = (request.data.get("review_note") or "").strip()
        row.is_paid = decision == "APPROVED"
        row.save(update_fields=["status", "reviewed_by", "reviewed_at", "review_note", "is_paid"])
        return Response({"detail": f"Leave {decision.lower()}."})


class PayrollAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        rows = PayrollRecord.objects.select_related("employee__user")
        if _role(request.user, "ADMIN"):
            if request.query_params.get("month"):
                rows = rows.filter(payroll_month=_month(request.query_params["month"]))
        else:
            rows = rows.filter(employee__user=request.user)
        return Response({"payroll": [{"id": r.id, "employee": r.employee.employee_id, "employee_name": r.employee.user.get_full_name(), "month": r.payroll_month, "base_salary": r.base_salary, "late_penalty": r.late_penalty, "half_day_deduction": r.half_day_deduction, "absence_deduction": r.absence_deduction, "overtime_hours": r.overtime_hours, "overtime_amount": r.overtime_amount, "rent_incentive": r.rent_incentive, "sale_incentive": r.sale_incentive, "net_salary": r.net_salary, "status": r.status, "snapshot": r.calculation_snapshot} for r in rows[:1000]]})

    @transaction.atomic
    def post(self, request):
        if not _role(request.user, "ADMIN"):
            return Response({"detail": "Only admin can generate payroll."}, status=403)
        try:
            month = _month(request.data.get("month", ""))
        except ValueError:
            return Response({"detail": "Month must be YYYY-MM."}, status=400)
        month_end = month.replace(day=calendar.monthrange(month.year, month.month)[1])
        if month_end >= timezone.localdate():
            return Response({"detail": "Payroll can be generated only after the month is complete."}, status=400)
        generated = 0
        for employee in EmployeeProfile.objects.filter(is_active=True, joining_date__lte=month_end):
            existing = PayrollRecord.objects.filter(employee=employee, payroll_month=month).first()
            if existing and existing.status != "DRAFT":
                continue
            result = calculate_payroll(employee, month)
            PayrollRecord.objects.update_or_create(employee=employee, payroll_month=month, defaults={
                "base_salary": result["base_salary"], "payable_base": result["payable_base"], "late_days": result["late_days"], "late_penalty": result["late_penalty"],
                "half_day_deduction": result["half_day_deduction"], "absence_deduction": result["absence_deduction"], "overtime_hours": result["overtime_hours"], "overtime_amount": result["overtime_amount"],
                "rent_incentive": result["rent_incentive"], "sale_incentive": result["sale_incentive"], "net_salary": result["net_salary"], "calculation_snapshot": result["snapshot"],
            })
            generated += 1
        return Response({"detail": "Payroll draft generated.", "records": generated})


class PayrollActionAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, payroll_id):
        if not _role(request.user, "ADMIN"):
            return Response({"detail": "Only admin can approve payroll."}, status=403)
        row = PayrollRecord.objects.filter(pk=payroll_id).first()
        if row is None:
            return Response({"detail": "Payroll record not found."}, status=404)
        action = request.data.get("action")
        if action == "APPROVE" and row.status == "DRAFT":
            row.status, row.approved_by, row.approved_at = "APPROVED", request.user, timezone.now()
            row.save(update_fields=["status", "approved_by", "approved_at", "updated_at"])
        elif action == "MARK_PAID" and row.status == "APPROVED":
            row.status, row.paid_at = "PAID", timezone.now()
            row.save(update_fields=["status", "paid_at", "updated_at"])
        else:
            return Response({"detail": "Invalid action for current payroll status."}, status=400)
        return Response({"detail": f"Payroll marked {row.status.lower()}."})


class PayrollExcelReportAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not _role(request.user, "ADMIN"):
            return Response({"detail": "Only admin can export salary register."}, status=403)
        try:
            month = _month(request.query_params.get("month", ""))
        except ValueError:
            return Response({"detail": "Month must be YYYY-MM."}, status=400)
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        rows = PayrollRecord.objects.filter(payroll_month=month).select_related("employee__user")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Salary Register"
        headers = ["Employee ID", "Employee", "Base Salary", "Payable Base", "Late Days", "Late Penalty", "Half-day Deduction", "Absence Deduction", "OT Hours", "OT Amount", "Rent Incentive", "Sale Incentive", "Other Earnings", "Other Deductions", "Net Salary", "Status"]
        sheet.append([f"ARI SMART RO — Salary Register {month:%B %Y}"])
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        sheet["A1"].fill = PatternFill("solid", fgColor="07315E")
        sheet["A1"].alignment = Alignment(horizontal="center")
        sheet.append(headers)
        for cell in sheet[2]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0868D7")
        for row in rows:
            sheet.append([row.employee.employee_id, row.employee.user.get_full_name(), row.base_salary, row.payable_base, row.late_days, row.late_penalty, row.half_day_deduction, row.absence_deduction, row.overtime_hours, row.overtime_amount, row.rent_incentive, row.sale_incentive, row.other_earnings, row.other_deductions, row.net_salary, row.status])
        sheet.freeze_panes = "A3"
        sheet.auto_filter.ref = f"A2:P{max(2, sheet.max_row)}"
        for column_index in range(1, sheet.max_column + 1):
            letter = get_column_letter(column_index)
            values = (
                sheet.cell(row=row_index, column=column_index).value
                for row_index in range(1, sheet.max_row + 1)
            )
            sheet.column_dimensions[letter].width = min(
                28,
                max(12, max(len(str(value or "")) for value in values) + 2),
            )
        stream = BytesIO()
        workbook.save(stream)
        response = HttpResponse(stream.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="ARI_Salary_Register_{month:%Y_%m}.xlsx"'
        return response
