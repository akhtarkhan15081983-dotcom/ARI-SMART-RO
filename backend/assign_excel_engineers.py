import os
import django
from collections import Counter

from django.db import transaction
from openpyxl import load_workbook

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
django.setup()

from customers.models import Customer
from employees.models import EmployeeProfile


EXCEL_FILE = "customers_master.xlsx"

TARGET = {
    "RAJKUMAR",
    "HASNAIN",
    "RAMA",
    "RUPESH",
}


# ============================================================
# FIND ENGINEERS
# ============================================================

employees = {}

for profile in EmployeeProfile.objects.select_related("user").filter(
    designation="ENGINEER",
    is_active=True,
):

    full_name = profile.user.get_full_name().strip().upper()

    if full_name == "RAJKUMAR KUSHWAH":
        employees["RAJKUMAR"] = profile

    elif full_name == "HUSNAIN GAZI":
        employees["HASNAIN"] = profile

    elif full_name == "RAMA AASRE":
        employees["RAMA"] = profile

    elif full_name == "RUPESH BHAGHEL":
        employees["RUPESH"] = profile


print()
print("=" * 70)
print("ENGINEERS")
print("=" * 70)

for name in sorted(TARGET):

    profile = employees.get(name)

    if profile:
        print(
            name,
            "->",
            profile.user.get_full_name(),
            "| Profile ID:",
            profile.id,
        )
    else:
        print(name, "-> NOT FOUND")


missing_engineers = TARGET - set(employees.keys())

if missing_engineers:

    print()
    print("ERROR: ENGINEER PROFILE MISSING:")
    for name in sorted(missing_engineers):
        print(" -", name)

    print()
    print("NO DATABASE RECORDS WERE CHANGED.")
    raise SystemExit(1)


# ============================================================
# LOAD DATABASE CUSTOMERS
# ============================================================

db_old = {
    str(customer.old_card_number).strip(): customer
    for customer in Customer.objects.exclude(old_card_number="")
}


# ============================================================
# READ EXCEL
# ============================================================

workbook = load_workbook(
    EXCEL_FILE,
    read_only=True,
    data_only=True,
)

sheet = workbook.active

rows = sheet.iter_rows(values_only=True)

# First row
next(rows)

# Second row = actual headers
headers = list(next(rows))

old_i = headers.index("card number")
name_i = headers.index("CUSTUMER NAME")
emp_i = headers.index("employee")


# ============================================================
# BUILD ASSIGNMENT LIST
# ============================================================

assignment_list = []

matched = Counter()
unmatched = Counter()


for excel_row, row in enumerate(rows, start=3):

    employee = row[emp_i]

    if employee is None:
        continue

    employee = str(employee).strip().upper()

    if employee not in TARGET:
        continue

    old_card = row[old_i]

    if old_card is None:
        old_card = ""
    else:
        old_card = str(old_card).strip()

    customer = db_old.get(old_card)

    if customer is None:

        unmatched[employee] += 1

        continue

    profile = employees[employee]

    matched[employee] += 1

    assignment_list.append(
        (
            customer,
            profile,
            excel_row,
            employee,
            old_card,
        )
    )


# ============================================================
# SAFETY CHECK
# ============================================================

total_matched = len(assignment_list)
total_unmatched = sum(unmatched.values())


print()
print("=" * 70)
print("FINAL ASSIGNMENT CHECK")
print("=" * 70)

for employee in sorted(TARGET):

    print(
        employee,
        "| Matched:",
        matched[employee],
        "| Unmatched:",
        unmatched[employee],
    )

print()
print("TOTAL MATCHED  :", total_matched)
print("TOTAL UNMATCHED:", total_unmatched)


# We expect exactly 778 matched customers and 57 unmatched.
if total_matched != 778:
    print()
    print("ERROR: Expected 778 matched customers.")
    print("Found:", total_matched)
    print()
    print("NO DATABASE RECORDS WERE CHANGED.")
    raise SystemExit(1)


if total_unmatched != 57:
    print()
    print("ERROR: Expected 57 unmatched customers.")
    print("Found:", total_unmatched)
    print()
    print("NO DATABASE RECORDS WERE CHANGED.")
    raise SystemExit(1)


# ============================================================
# ACTUAL DATABASE UPDATE
# ============================================================

print()
print("=" * 70)
print("STARTING DATABASE ASSIGNMENT")
print("=" * 70)

changed = 0
already_correct = 0


with transaction.atomic():

    for (
        customer,
        profile,
        excel_row,
        employee,
        old_card,
    ) in assignment_list:

        if customer.assigned_engineer_id == profile.id:

            already_correct += 1

            continue

        customer.assigned_engineer = profile

        customer.save(
            update_fields=[
                "assigned_engineer",
            ]
        )

        changed += 1


print()
print("=" * 70)
print("ASSIGNMENT COMPLETE")
print("=" * 70)

print("Matched customers :", total_matched)
print("Database changed  :", changed)
print("Already correct   :", already_correct)
print("Unmatched/skipped :", total_unmatched)

print()
print("57 unmatched customers were NOT modified.")
print("Rent history was NOT modified.")
print("Payment data was NOT modified.")
print("Customer details were NOT modified.")

print()
print("DATABASE ASSIGNMENT SUCCESSFUL.")