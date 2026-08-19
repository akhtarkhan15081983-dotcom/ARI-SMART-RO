import os
import django
from collections import Counter
from openpyxl import load_workbook

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
django.setup()

from customers.models import Customer


EXCEL_FILE = "customers_master.xlsx"

TARGET = {
    "RAJKUMAR",
    "HASNAIN",
    "RAMA",
    "RUPESH",
}


workbook = load_workbook(
    EXCEL_FILE,
    read_only=True,
    data_only=True,
)

sheet = workbook.active

rows = sheet.iter_rows(values_only=True)

# Row 1 = totals/header area
next(rows)

# Row 2 = actual headers
headers = list(next(rows))

old_i = headers.index("card number")
name_i = headers.index("CUSTUMER NAME")
emp_i = headers.index("employee")
phone_i = headers.index("Contect No.")


db_old = {
    str(customer.old_card_number).strip(): customer
    for customer in Customer.objects.exclude(
        old_card_number=""
    )
}


total = Counter()
matched = Counter()
unmatched = Counter()


unmatched_rows = []


for excel_row, row in enumerate(rows, start=3):

    employee = row[emp_i]

    if employee is None:
        continue

    employee = str(employee).strip().upper()

    if employee not in TARGET:
        continue

    total[employee] += 1

    old_card = row[old_i]

    if old_card is None:
        old_card = ""
    else:
        old_card = str(old_card).strip()

    if old_card in db_old:

        matched[employee] += 1

    else:

        unmatched[employee] += 1

        unmatched_rows.append(
            (
                excel_row,
                employee,
                old_card,
                row[name_i],
                row[phone_i],
            )
        )


print()
print("=" * 70)
print("CUSTOMER ASSIGNMENT MATCH CHECK")
print("=" * 70)

print()

for employee in sorted(TARGET):

    print(
        employee,
        "| Excel:",
        total[employee],
        "| Matched:",
        matched[employee],
        "| Unmatched:",
        unmatched[employee],
    )


print()

print("TOTAL EXCEL   :", sum(total.values()))
print("TOTAL MATCHED :", sum(matched.values()))
print("TOTAL UNMATCHED:", sum(unmatched.values()))

print()
print("=" * 70)
print("UNMATCHED TARGET CUSTOMERS")
print("=" * 70)

for item in unmatched_rows:

    excel_row, employee, old_card, name, phone = item

    print(
        "ROW:",
        excel_row,
        "| EMP:",
        employee,
        "| OLD:",
        old_card,
        "| NAME:",
        name,
        "| PHONE:",
        phone,
    )

print()
print("NO DATABASE RECORDS WERE CHANGED.")